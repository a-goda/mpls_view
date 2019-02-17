import re
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment, Font, colors, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
import os
import fnmatch
from os.path import isfile, join
import time
import sqlite3
import ipaddress
from string import ascii_uppercase
# for profiler
import cProfile
import pstats
import io
import socket


rgx_host = r'(?P<hostname>[^\s]+)#\s*(sh[ow]{0,2}\s*run.*?)?' \
           r'(sh[ow]{0,2}\s*cdp\s*ne[ighbors]{0,7}\s*(d[etails]{0,6})?)?' \
           r'(sh[ow]{0,2}\s*vlan)?'
rgx_show_run_sec = r'(?P<hostname>.+?)#\s*sh[ow]{0,2}\s+run[nig]{0,4}.*'
rgx_cdp_nei_sec = r'(?P<hostname>.+?)#\s*sh[ow]{0,2}\s*cdp\s*ne[ighbors]{0,7}\s*(d[etails]{0,6})?'
rgx_show_vlan_sec = r'(?P<hostname>.+?)#\s*sh[ow]{0,2}\s*vlan'
rgx_show_inventory = r'(?P<hostname>.+?)#\s*sh[ow]{0,2}\s*in[vtory]{0,5}'

rgx_ios_ip_vrf = r'\s*ip\s*vrf\s*(?P<vrf_name>.+)'
rgx_ios_vrf_def = r'\s*vrf\s*definition\s*(?P<vrf_name>.+)'

rgx_xr_vrf = r'\s*vrf\s+(?P<vrf_name>.+)\s*'
rgx_interface_vrf_fwd = r'\s*(ip\s+)?vrf\s+(forwarding\s+)?(?P<vrf_name>.+)'
rgx_vrf_fwd = r'\s*ip\s*vrf\s*forwarding\s*(?P<vrf_name>.+)'
rgx_nx_vrf_fwd = r'\s+vrf\s+member\s+(?P<vrf_name>\S+)'

rgx_rd = r'\s*rd\s*(?P<rd>.+)'
rgx_add_family = r'\s*address.family\s*ip(?P<add_family>v[64])\s*(?P<transport_type>.*)'
rgx_ios_exp_rt = r'\s*route-target\s*export\s*(?P<rt_export>.+)'
rgx_ios_imp_rt = r'\s*route-target\s*import\s*(?P<rt_import>.+)'

rgx_xr_exp_rt = r'\s*export\s*route-target\s*$'
rgx_xr_imp_rt = r'\s*import\s*route-target\s*$'
rgx_xr_rt = r'[0-9:]+$'

rgx_description = r'\s*description\s*(?P<description>.+)'
rgx_import_map = r'\s*import\s+map\s+(?P<map>.+)'
rgx_export_map = r'\s*export\s+map\s+(?P<map>.+)'

rgx_interface = r'\s*[iI]nterface\s+(?P<type>[^0-9]+)(?P<number>[0-9a-zA-Z\\/]+(\.(?P<sub>\d+))?)'
# rgx_inter_type_num = r'(?P<type>[^0-9]+)(?P<number>[0-9\\/]+)'
rgx_tunnel_interface = r'\s*[iI]nterface\s+[tT]unnel(?P<number>.+)'

rgx_ip_add = r'\s*ip(?P<ver>v\d)?(\s+address)?\s+(?P<ip_add>(?P<ip>\d+\.\d+\.\d+\.\d+)' \
             r'(\s*(?P<subnet>\d+\.\d+\.\d+\.\d+))?)(\s*(?P<sec>[sS]econdary))?'

rgx_nx_ip_add = r'\s+ip\s+address\s+(?P<ip_add>(?P<ip>\d+\.\d+\.\d+\.\d+))' \
                r'/(?P<subnet>\d+)?(\s*(?P<sec>[sS]econdary))?'

rgx_ipv6_add = r'\s*ipv6\s+address\s+(?P<ip_add>[0-9:]+(/(?P<subnet>\d+))?)\s*(?P<type>\S+)?'

rgx_ios_static_route = r'\s*ip(?P<add_fam>v6)?\s+route\s+(vrf\s+(?P<vrf>.+?)\s.+)?' \
                   r'(?P<sub>(?P<net>\d+\.\d+\.\d+\.\d+)\s+(?P<mask>\d+\.\d+\.\d+\.\d+))\s+' \
                   r'(?P<next_hop>((?P<ip_next_hop>\d+\.\d+\.\d+\.\d+)\s*)?' \
                   r'((?P<int_next_hop>(?P<type>[^0-9]+)(?P<number>[0-9\\/]+)(?P<sub_int>\.\d+)?)\s*)?)' \
                   r'((?P<ad>\d+\s*))?(name\s+(?P<name>.+))?'

rgx_xr_router_static = r'router\s*static\s*$'

rgx_xr_static_route_entry = r'\s*(?P<sub>(?P<net>\d+\.\d+\.\d+\.\d+)/(?P<mask>\d{1,2}))\s+(vrf\s+(?P<vrf>.+?)\s+)?' \
                            r'((?P<int_next_hop>(?P<type>[^0-9]+)(?P<number>[0-9\\/]+)(?P<sub_int>\.\d+)?)\s*)?' \
                            r'((?P<ip_next_hop>\d+\.\d+\.\d+\.\d+)\s*)?((?P<ad>\d+\s*))?' \
                            r'(description\s+(?P<description>.+))?'

dark_blue = '0000b3'
blue_white = 'e6e6e6'
gray_border = 'a6a6a6'
# 'f2f2f2'
# 'e6f0ff'

db_con = None
db_cur = None

banded = False

print_ignore_break = False

# VRF sheet columns variables
col_hostname = ''
col_vrf = ''
col_rd = ''
col_export_rt = ''
col_imports_pe = ''
col_imports_rts = ''
col_exported_to_pe = ''
col_exported_to_rts = ''

# Interface sheet columns variables
col_int_interface_name = ''
col_int_desc = ''
col_int_addr = ''

col_route_next_hop = ''
col_route_count = ''

add_family_v4_id = 1
add_family_v6_id = 2


def profile(fnc):
    """
    A decorator that uses cProfile to profile a function

    :param fnc: profiling function
    :return: retun_values of inner function
    """
    def inner(*args, **kwargs):
        pr = cProfile.Profile()
        pr.enable()
        return_vals = fnc(*args, **kwargs)
        pr.disable()
        s = io.StringIO()
        sortby = 'cumulative'
        ps = pstats.Stats(pr, stream=s).sort_stats(sortby)
        ps.print_stats()
        print(s.getvalue())
        return return_vals
    return inner


def prepare_workbooks_styles(wb_list):
    for wb in wb_list:
        if 'Wrap_center_left' not in wb.named_styles:
            cell_style = NamedStyle(name='Wrap_center_left')
            cell_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            wb.add_named_style(cell_style)

        if 'banded_row' not in wb.named_styles:
            cell_style = NamedStyle(name='banded_row')
            cell_style.fill = PatternFill(start_color=blue_white, end_color=blue_white, fill_type='solid')
            cell_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell_style.border = Border(right=Side(border_style='thin', color=gray_border),
                                       left=Side(border_style='thin', color=gray_border),
                                       top=Side(border_style='thin', color=gray_border),
                                       bottom=Side(border_style='thin', color=gray_border))

            wb.add_named_style(cell_style)

        if 'center_left' not in wb.named_styles:
            cell_style = NamedStyle(name='center_left')
            cell_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell_style.border = Border(right=Side(border_style='thin', color=gray_border),
                                       left=Side(border_style='thin', color=gray_border),
                                       top=Side(border_style='thin', color=gray_border),
                                       bottom=Side(border_style='thin', color=gray_border))
            wb.add_named_style(cell_style)

        if 'center_center' not in wb.named_styles:
            cell_style = NamedStyle(name='center_center')
            cell_style.alignment = Alignment(horizontal='center', vertical='center')
            wb.add_named_style(cell_style)

        if 'center_center_bold' not in wb.named_styles:
            cell_style = NamedStyle(name='center_center_bold')
            cell_style.alignment = Alignment(horizontal='center', vertical='center')
            cell_style.font = Font(bold=True, color=colors.WHITE)
            cell_style.border = Border(right=Side(border_style='thin', color=colors.WHITE),
                                       top=Side(border_style='thin', color=colors.WHITE))
            cell_style.fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type='solid')
            wb.add_named_style(cell_style)

        if 'center_center_bold_12' not in wb.named_styles:
            cell_style = NamedStyle(name='center_center_bold_12')
            cell_style.alignment = Alignment(horizontal='center', vertical='center')
            cell_style.font = Font(bold=True, size=12, color=colors.WHITE)
            cell_style.border = Border(right=Side(border_style='thin', color=colors.WHITE))
            cell_style.fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type='solid')
            wb.add_named_style(cell_style)


def prepare_one_row_sheet_header(wbas, heads_tuple, freeze_pane='A2'):
    letters = list(ascii_uppercase)[:len(heads_tuple)]

    for i, head in enumerate(heads_tuple):
        wbas[letters[i] + '1'] = head

    for cells in wbas['{0}1:{1}1'.format(letters[0], letters[-1])]:
        for cell in cells:
            cell.style = 'center_center_bold_12'

    wbas.freeze_panes = freeze_pane
    return 2


def write_one_sheet_row(wbas, value_tuple, row_num, notes='', col_limit=0):

    if not col_limit:
        for i, value in enumerate(value_tuple, 1):
            wbas.cell(row=row_num, column=i).value = value
        if notes:
            wbas.cell(row=row_num, column=len(value_tuple) + 1).value = notes

    else:
        for i, value in enumerate(value_tuple, 1):
            if i > col_limit:
                break
            wbas.cell(row=row_num, column=i).value = value
        if notes:
            wbas.cell(row=row_num, column=col_limit + 1).value = notes

    global banded

    if banded:
        for cells in wbas['A{0}:{1}{0}'.format(row_num, get_column_letter(wbas.max_column))]:
            for cell in cells:
                cell.style = 'banded_row'

        banded = False
    else:
        for cells in wbas['A{0}:{1}{0}'.format(row_num, get_column_letter(wbas.max_column))]:
            for cell in cells:
                cell.style = 'center_left'
        banded = True

    return row_num + 1


def find_max_rows(columns_list):
    num = 0
    for i, column in enumerate(columns_list):
        if i + 1 < len(columns_list):
            num = max(num,
                      len(column) if column is not None else 0,
                      len(columns_list[i + 1]) if columns_list[i + 1] is not None else 0)
        else:
            num = max(num, len(column) if column is not None else 0)
    # num = max(len(l1), len(l2), len(l3))
    if num > 0:
        return num
    else:
        return 1


def write_list(ws, start_row, column, data_list):
    for d in data_list:
        ws.cell(row=start_row, column=column).value = d[0]
        start_row = start_row + 1


def write_export_to_import_from_columns(ws, start_row, start_column, data_list):
    # for tup_list in data_list:
    row_num = start_row
    col_num = start_column
    for d in data_list:
        ws.cell(row=row_num, column=col_num).value = d[0]
        ws.cell(row=row_num, column=col_num + 1).value = d[1] + "({0})".format(d[3])
        row_num = row_num + 1


def auto_column_width_vrf(ws, start_row=3, max_width=55, is_vrf_sheet=True):
    for col in ws.columns:
        max_length = 1
        column = col[0].column  # Get the column name

        i = start_row
        for cell in col:
            if cell.value is not None and len(str(cell.value)) > max_length:
                # str: as the column may of integer type
                max_length = len(str(cell.value))

            i = i + 1

        if is_vrf_sheet:
            if max_length < 3:
                # for empty columns
                adjusted_width = 6
            elif max_length > max_width:
                adjusted_width = max_width
            else:
                adjusted_width = max_length + 2
        else:
            adjusted_width = max_length + 2

        ws.column_dimensions[column].width = adjusted_width


def str_to_int(text):
    return int(text) if text.isdigit() else text


def natural_sort(text):
    return [str_to_int(c) for c in re.split(r'(\d+)', text)]


def prepare_vrf_sheet_header(wbas, start_row=1, freeze=True):
    col_head_hostname = 'A{0}:A{1}'.format(start_row, start_row + 1)
    col_head_vrf = 'B{0}:D{0}'.format(start_row)

    col_head_imports = 'E{0}:F{0}'.format(start_row)
    col_head_exported_to = 'G{0}:H{0}'.format(start_row)
    col_head_interface = 'I{0}:L{0}'.format(start_row)
    col_head_routes = 'M{0}:N{0}'.format(start_row)

    col_coord_hostname = 'A{0}'.format(start_row)
    col_coord_vrf = 'B{0}'.format(start_row + 1)
    col_coord_rd = 'C{0}'.format(start_row + 1)
    col_coord_export_rt = 'D{0}'.format(start_row + 1)
    col_coord_imports_pe = 'E{0}'.format(start_row + 1)
    col_coord_imports_rts = 'F{0}'.format(start_row + 1)
    col_coord_exported_to_pe = 'G{0}'.format(start_row + 1)
    col_coord_exported_to_rts = 'H{0}'.format(start_row + 1)
    col_coord_int_name = 'I{0}'.format(start_row + 1)
    col_coord_int_ips = 'J{0}'.format(start_row + 1)
    col_coord_int_desc = 'K{0}'.format(start_row + 1)
    col_coord_int_note = 'L{0}'.format(start_row + 1)
    col_coord_route_next_hop = 'M{0}'.format(start_row + 1)
    col_coord_route_count = 'N{0}'.format(start_row + 1)

    wbas.merge_cells(col_head_vrf)
    wbas[col_head_vrf[:col_head_vrf.index(':')]] = "VRF"
    wbas.merge_cells(col_head_imports)
    wbas[col_head_imports[:col_head_imports.index(':')]] = "Import From"
    wbas.merge_cells(col_head_exported_to)
    wbas[col_head_exported_to[:col_head_exported_to.index(':')]] = "Imported By"
    wbas.merge_cells(col_head_interface)
    wbas[col_head_interface[:col_head_interface.index(':')]] = "L3 Interfaces"
    wbas.merge_cells(col_head_routes)
    wbas[col_head_routes[:col_head_routes.index(':')]] = "Static Routes Summary"

    wbas.merge_cells(col_head_hostname)
    wbas[col_coord_hostname] = 'PE Hostname'
    wbas[col_coord_vrf] = "VRF Name"
    wbas[col_coord_rd] = 'RD'

    wbas[col_coord_export_rt] = 'Exported RT'
    wbas[col_coord_imports_pe] = 'PE (Alias)'
    wbas[col_coord_imports_rts] = 'VRF Name(RT)'
    wbas[col_coord_exported_to_pe] = 'PE (Alias)'
    wbas[col_coord_exported_to_rts] = 'VRF Name(RT)'

    wbas[col_coord_int_name] = 'Int. Name [VLAN name](Status)(VLAN_ID)'
    wbas[col_coord_int_ips] = 'IP Addresses (Primary/VIP)'
    wbas[col_coord_int_desc] = 'Int. Description'
    wbas[col_coord_int_note] = 'Notes'

    wbas[col_coord_route_next_hop] = 'Next Hop (IP/Int) ->(To_VRF)'
    wbas[col_coord_route_count] = 'Count'

    """
    Auto column number generation, data write to Excel file method uses the below variables to select 
    the appropriate column to save data.

    Overall goal: changing column order above, reflect automatically on all preceding methods
    """
    global col_hostname, col_vrf, col_rd, col_export_rt, col_imports_pe, col_imports_rts, col_exported_to_pe, \
        col_exported_to_rts, col_int_interface_name, col_int_addr, col_int_desc, col_route_next_hop, col_route_count

    col_hostname = column_index_from_string(re.search(r'(\D+)', col_coord_hostname)[0])
    col_vrf = column_index_from_string(re.search(r'(\D+)', col_coord_vrf)[0])
    col_rd = column_index_from_string(re.search(r'(\D+)', col_coord_rd)[0])
    col_export_rt = column_index_from_string(re.search(r'(\D+)', col_coord_export_rt)[0])
    col_imports_pe = column_index_from_string(re.search(r'(\D+)', col_coord_imports_pe)[0])
    col_imports_rts = column_index_from_string(re.search(r'(\D+)', col_coord_imports_rts)[0])
    col_exported_to_pe = column_index_from_string(re.search(r'(\D+)', col_coord_exported_to_pe)[0])
    col_exported_to_rts = column_index_from_string(re.search(r'(\D+)', col_coord_exported_to_rts)[0])

    col_int_interface_name = column_index_from_string(re.search(r'(\D+)', col_coord_int_name)[0])
    col_int_addr = column_index_from_string(re.search(r'(\D+)', col_coord_int_ips)[0])
    col_int_desc = column_index_from_string(re.search(r'(\D+)', col_coord_int_desc)[0])
    col_route_next_hop = column_index_from_string(re.search(r'(\D+)', col_coord_route_next_hop)[0])
    col_route_count = column_index_from_string(re.search(r'(\D+)', col_coord_route_count)[0])

    for cells in wbas['A{1}:{0}{1}'.format(get_column_letter(wbas.max_column), start_row)]:
        for cell in cells:
            cell.style = 'center_center_bold_12'

    for cells in wbas['A{1}:{0}{1}'.format(get_column_letter(wbas.max_column), start_row + 1)]:
        for cell in cells:
            cell.style = 'center_center_bold'

    if freeze:
        wbas.freeze_panes = 'A' + str(wbas.max_row + 1)

    # Return next usable row, after header
    return wbas.max_row + 1


def write_one_summary_row(wbas, value_tuple, row_num):

    for i, value in enumerate(value_tuple, 1):
        if isinstance(value, tuple):
            wbas.cell(row=row_num, column=i).hyperlink = value[0]
            wbas.cell(row=row_num, column=i).value = value[1]
        else:
            wbas.cell(row=row_num, column=i).value = value

    global banded

    if banded:
        for cells in wbas['A{0}:{1}{0}'.format(row_num, get_column_letter(wbas.max_column))]:
            for cell in cells:
                cell.style = 'banded_row'

        banded = False
    else:
        for cells in wbas['A{0}:{1}{0}'.format(row_num, get_column_letter(wbas.max_column))]:
            for cell in cells:
                cell.style = 'center_left'
        banded = True


def write_vrf_summary(sum_ws, sheets):
    """
    Write the summary sheet
    :param sum_ws: workbook sheet object
    :param sheets: [ [sheetname, core_name, starting position, vrf count, interfaces count] ]
    """
    prepare_one_row_sheet_header(sum_ws, ('#', 'Core', 'Number of VRFs', 'Number of L3 Interfaces'), freeze_pane='B2')
    global banded

    banded = False

    for i, sheet in enumerate(sheets, 2):
        sheet_name = sheet[0]
        core_name = sheet[1]
        start_row = sheet[2] if sheet[2] else 1
        vrf_count = sheet[3]
        l3_count = sheet[4]

        write_one_summary_row(sum_ws, (i - 1, ('#{0}!A{1}'.format(sheet_name, start_row), core_name),
                                       vrf_count, l3_count), i)

    auto_column_width_vrf(sum_ws, start_row=1, is_vrf_sheet=False)


def back_to_summary_link(ws, row_num, summary_back_row, column=(5, 13), sheet_name='Summary'):
    if isinstance(column, int):
        ws.cell(row=row_num, column=column).hyperlink = '#{0}!B{1}'.format(sheet_name, summary_back_row)
        ws.cell(row=row_num, column=column).value = 'Back To Summary'
        ws.cell(row=row_num, column=column).style = 'Hyperlink'
    elif isinstance(column, tuple):
        for col in column:
            ws.cell(row=row_num, column=col).hyperlink = '#{0}!B{1}'.format(sheet_name, summary_back_row)
            ws.cell(row=row_num, column=col).value = 'Back To Summary'
            ws.cell(row=row_num, column=col).style = 'Hyperlink'


def save_vrfs_to_excel(wb):
    appliances_data = do_query(
        """ select site.name, appliance.hostname, appliance.app_id, log_file.importance 
            from appliance 
            join site on appliance.site_id = site.site_id
            join log_file on appliance.app_id = log_file.app_id;
        """)
    global banded
    branches = False
    branches_wbs = wb.active

    sum_ws = wb.create_sheet('Summary', 0)
    sheets = []

    for i, switch in enumerate(appliances_data):
        filename = switch[0]
        hostname = switch[1]
        sw_id = switch[2]
        importance = switch[3]

        # Makes banded row start the same for all sheets
        # first row is not banded
        banded = False

        if importance:
            # if filename has a starting number
            ws = wb.create_sheet(filename, i + 1)
            next_row = prepare_vrf_sheet_header(ws, start_row=2)
            sheets.append([filename, filename, 0])
            back_to_summary_link(ws, row_num=1, summary_back_row=len(sheets) + 1)

        else:
            # normal filename without a starting number
            last_row = 1
            if not branches:
                ws = wb.create_sheet('Branches', i + 1)
                ws['A{0}'.format(last_row)] = 'Core Name:'
                ws['B{0}'.format(last_row)] = filename

                next_row = prepare_vrf_sheet_header(ws, start_row=2, freeze=False)
                branches_wbs = ws
                branches = True
                sheets.append(['Branches', filename, 1])
                back_to_summary_link(ws, row_num=last_row, summary_back_row=len(sheets) + 1)

            else:
                ws = branches_wbs
                last_row = ws.max_row + 3
                ws['A{0}'.format(last_row)] = 'Core Name:'
                ws['B{0}'.format(last_row)] = filename

                next_row = prepare_vrf_sheet_header(ws, start_row=last_row + 1, freeze=False)
                sheets.append(['Branches', filename, last_row])
                back_to_summary_link(ws, row_num=last_row, summary_back_row=len(sheets) + 1)

            for cells in ws['A{0}:B{0}'.format(last_row)]:
                for cell in cells:
                    cell.style = 'Title'

        vrfs = do_query(
            """ select vrf.name, vrf.rd, vrf.vrf_id from vrf where vrf.app_id = {0}
            """.format(sw_id))
        vrfs.sort(key=lambda vrf_rd: sort_vrf_rds(vrf_rd[1]))
        sheets[-1].extend((len(vrfs), 0))

        for vrf in vrfs:
            vrf_name = vrf[0]
            rd = vrf[1]
            vrf_id = vrf[2]
            exported_to = get_exported_to(vrf_id)
            import_from = get_import_from_vrfs(vrf_id)
            interfaces = find_vrf_interfaces(sw_id, vrf_id, include_default_vrf=True)
            sheets[-1][4] = sheets[-1][4] + len(interfaces)

            routes_summary = find_static_routes_summary(sw_id, vrf_id, include_null_values=True)

            vrf_exported_rt = do_query("""select rt_export from export_rt where vrf_id = {0}""".format(vrf_id))

            next_row = write_vrf_sheet_row(ws, hostname, vrf_name, rd,
                                           [vrf_exported_rt, exported_to, import_from, interfaces, routes_summary],
                                           next_row)
        auto_column_width_vrf(ws)

    write_vrf_summary(sum_ws, sheets)


def sort_vrf_rds(rd):
    if not rd:
        return 0
    else:
        return int(rd[rd.index(':') + 1:])


def sort_interface_num_key(item):
    if '/' not in item:
        return int(item)
    else:
        return int(item[:item.index('/')])


def sort_ip_next_hop_key(item):
    # item = item[3]
    if re.match(r'[A-Za-z]+(?P<num>\d+)', item):
        return re.match(r'[A-Za-z]+(?P<num>\d+)', item)['num']
    elif ',' not in item:
        return socket.inet_aton(item[:item.index('/')])
    else:
        item = item.split(',')[0].strip()
        return socket.inet_aton(item[:item.index('/')])


def write_vrf_interfaces(wbas, start_row, start_column, data_list):
    """
    Receive a list of interfaces' details
    (vrf_name, vrf_export_rt(s), int_name_type, ip_address(es), int_description, sw_id, int_type, int_number, int_id)
    Append to the active workbook sheet starting from start_row and start_column
    :param wbas: active workbook sheet
    :param start_row: starting sheet row number
    :param start_column: starting sheet column number
    :param data_list: list of tuples
    :return:
    """
    row_num = start_row
    col_num = start_column

    for d in data_list:

        wbas.cell(row=row_num, column=col_num).value = d[2]
        wbas.cell(row=row_num, column=col_num + 1).value = d[3]
        wbas.cell(row=row_num, column=col_num + 2).value = d[4]

        if d[6] == 'vlan':
            qr = do_query('select vlan_id, exist from vlan where vlan_no={0} and app_id={1}'.format(d[7], d[5]), True)
            if not qr:
                wbas.cell(row=row_num, column=col_num + 3).value = 'No L2 VLAN ID is found'
            if qr and not qr[1]:
                int_qr = do_query("""
                select group_concat(interface.type||interface.number, ', ')
                from interface
                join int_vlan on interface.int_id = int_vlan.int_id
                join vlan on int_vlan.vlan_id = vlan.vlan_id
                where vlan.exist != 1 and int_vlan.vlan_id = {0} and interface.app_id = {1};
                """.format(qr[0], d[5]), True)
                wbas.cell(row=row_num, column=col_num + 3).value \
                    = 'No L2 VLAN ID is found' + \
                      ', Tagged/Access on Interfaces ({0})'.format(int_qr[0]) if int_qr else ''

        elif d[6] == 'tunnel':
            # find which appliance the tunnel destination IP is configured on, and on which interface
            # staring from finding tunnel destination IP, source IP/Int
            tunn_src_dst = do_query("""
            select ip_address.address, tunnel_int.source_int, tunnel_int.source_ip
            from interface
            join tunnel_int on interface.tunnel_id=tunnel_int.tunnel_id
            join ip_address on ip_address.ip_id = tunnel_int.dest_ip
            join appliance on interface.app_id = appliance.app_id
            join site on appliance.site_id = site.site_id
            where interface.int_id = {0};
             """.format(d[8]), True)

            if tunn_src_dst:
                if tunn_src_dst[1]:
                    src = do_query("""
                    select
                    case
                    when interface.type='vlan' then upper(interface.type)||interface.number
                    else (interface.type||interface.number)
                    end as interface  
                    from interface
                    where int_id = {0};
                    """.format(tunn_src_dst[1]), True)
                else:
                    src_ip = do_query("select address from ip_address where ip_id = {0}".format(tunn_src_dst[2]), True)
                    src_ip = src_ip[0][:src_ip[0].index('/')]
                    src = do_query("""
                    select
                    case
                        when interface.type='vlan' then upper(interface.type)||interface.number
                        else (interface.type||interface.number)
                        end as interface
                    from interface 
                    where interface.int_id = (
                    select ip_address.int_id 
                    from ip_address 
                    where address like "{0}/%" and (int_id is not null or int_id != ""));
                    """.format(src_ip), True)
                    if src:
                        src = [src_ip + ' ({0})'.format(src[0])]
                    else:
                        src = [src_ip + ' (N/A)']
                dst_ip = tunn_src_dst[0][:tunn_src_dst[0].index('/')]
                qr = do_query("""
                select site.name, 
                case
                when interface.type='vlan' then upper(interface.type)||interface.number
                else (interface.type||interface.number)
                end as interface
                from interface
                join appliance on interface.app_id = appliance.app_id
                join site on appliance.site_id = site.site_id
                join ip_address on ip_address.int_id = interface.int_id
                where  ip_address.address like "{0}%"
                """.format(dst_ip), True)
                if qr:
                    wbas.cell(row=row_num, column=col_num + 3).value = 'Src: {0}, Dest: {1} ({2}, {3})'.format(
                        src[0], dst_ip, qr[0], qr[1])


        row_num = row_num + 1


def write_extra_columns(wbas, start_row, start_column, data_list):
    row_num = start_row

    for d in data_list:
        for col_num, dd in enumerate(d, start_column):
            wbas.cell(row=row_num, column=col_num).value = dd

        row_num = row_num + 1


def write_vrf_sheet_row(wbas, hostname, vrf_name, rd, vrf_data_list, start_row):
    """
    Append an Excel sheet row for a VRF, calculates the maximum column heights for all columns, and then
    merge cells for hostname, VRF name and RD columns based on it
    :param wbas: active workbook sheet
    :param hostname: appliance hostname
    :param vrf_name: vrf name
    :param rd: vrf rd
    :param vrf_data_list: [vrf_exported_rt, exported_to, import_from, interfaces]
    :param start_row: the starting row where the data has to be appended to
    :return: next usable row, after the last row appended
    """
    # (ws, hostname, vrf_name, rd, [vrf_exported_rt, exported_to, import_from, interfaces], next_row)
    # Adding max number of rows (including start row) to the start row will count start row twice, then -1
    vrf_exported_rt = vrf_data_list[0]
    exported_to = vrf_data_list[1]
    import_from = vrf_data_list[2]
    interfaces = vrf_data_list[3]
    static_routes_summary = vrf_data_list[4]

    end_row = find_max_rows(
        (vrf_exported_rt, exported_to, import_from, interfaces, static_routes_summary)) + start_row - 1

    wbas.merge_cells(start_row=start_row, start_column=col_hostname, end_row=end_row, end_column=col_hostname)
    wbas.cell(row=start_row, column=col_hostname).value = hostname

    wbas.merge_cells(start_row=start_row, start_column=col_vrf, end_row=end_row, end_column=col_vrf)
    wbas.cell(row=start_row, column=col_vrf).value = vrf_name

    wbas.merge_cells(start_row=start_row, start_column=col_rd, end_row=end_row, end_column=col_rd)
    wbas.cell(row=start_row, column=col_rd).value = rd

    if len(vrf_exported_rt) == 1:
        wbas.merge_cells(start_row=start_row, start_column=col_export_rt, end_row=end_row, end_column=col_export_rt)
        wbas.cell(row=start_row, column=col_export_rt).value = vrf_exported_rt[0][0]
    elif not vrf_exported_rt:
        wbas.merge_cells(start_row=start_row, start_column=col_export_rt, end_row=end_row, end_column=col_export_rt)
    else:
        write_list(wbas, start_row, col_export_rt, vrf_exported_rt)

    if not exported_to:
        wbas.merge_cells(start_row=start_row, start_column=col_exported_to_pe,
                         end_row=end_row, end_column=col_exported_to_pe)
        wbas.merge_cells(start_row=start_row, start_column=col_exported_to_pe + 1,
                         end_row=end_row, end_column=col_exported_to_pe + 1)
    else:
        write_export_to_import_from_columns(
            wbas, start_row=start_row, start_column=col_exported_to_pe, data_list=exported_to)

    if not import_from:
        wbas.merge_cells(start_row=start_row, start_column=col_imports_pe,
                         end_row=end_row, end_column=col_imports_pe)
        wbas.merge_cells(start_row=start_row, start_column=col_imports_pe + 1,
                         end_row=end_row, end_column=col_imports_pe + 1)
    else:
        write_export_to_import_from_columns(wbas, start_row=start_row,
                                            start_column=col_imports_pe, data_list=import_from)

    if interfaces:
        write_vrf_interfaces(wbas, start_row=start_row, start_column=col_int_interface_name, data_list=interfaces)
    if static_routes_summary:
        write_extra_columns(wbas, start_row=start_row, start_column=col_route_next_hop, data_list=static_routes_summary)

    global banded

    if banded:
        for cells in wbas['A{0}:{1}{2}'.format(start_row, get_column_letter(wbas.max_column), end_row)]:
            for cell in cells:
                cell.style = 'banded_row'

        banded = False
    else:
        for cells in wbas['A{0}:{1}{2}'.format(start_row, get_column_letter(wbas.max_column), end_row)]:
            for cell in cells:
                cell.style = 'center_left'
        banded = True

    return end_row + 1


def save_int_to_excel(wb, headers):
    headers.append('Notes')

    switches_data = do_query(
        """ select site.name, appliance.app_id from appliance 
            join site on appliance.site_id = site.site_id;
        """)
    global banded

    for i, switch in enumerate(switches_data):
        filename = switch[0]
        sw_id = switch[1]

        ws = wb.create_sheet(filename, i)
        next_row = prepare_one_row_sheet_header(ws, headers)

        # Makes banded row start the same for all sheets
        # first row is not banded
        banded = False
        vrf_ids = do_query(
            """ select vrf.vrf_id from vrf where vrf.app_id = {0};
            """.format(sw_id))
        for vrf_id in vrf_ids:
            interfaces = find_vrf_interfaces(sw_id, vrf_id[0], include_default_vrf=True)

            for interface in interfaces:
                # interface = interface[:3]
                if re.match(r'VLAN(?P<id>\d+)', interface[2]):
                    vlan_id = do_query(
                        'select vlan_id from vlan '
                        'where vlan_no={0} and app_id={1}'.format(
                            re.match(r'VLAN(?P<id>\d+)', interface[2])['id'], sw_id), True)
                    if not vlan_id:
                        next_row = write_one_sheet_row(ws, interface, next_row, notes='No VLAN id found', col_limit=5)
                    else:
                        next_row = write_one_sheet_row(ws, interface, next_row, col_limit=5)
                else:
                    next_row = write_one_sheet_row(ws, interface, next_row, col_limit=5)

            auto_column_width_vrf(ws, start_row=1, is_vrf_sheet=False)


def create_db():
    """
    Create DB on memory, and return connection and cursor objects
    :return: db connection, db cursor
    """
    con = sqlite3.connect(':memory:')
    cursor = con.cursor()
    with open('db schema.sql', "r") as sf:
        cursor.executescript(sf.read())
        con.commit()

    return con, cursor


def db_close_dump(db_file):
    """
    Store memory database into SQL dump file, and close the current DB
    :param db_file: SQL output dump file
    """

    with open(db_file, 'w') as f:
        for line in db_con.iterdump():
            f.write('{0}\n'.format(line))

    db_con.close()


def prepare_string_with_quotes(data_set, quotes=None):
    """
    Prepare input data for DB queries, with quotes if necessary
    :param data_set: entry(s) to be prepared
    :param quotes: if its required to quotes data should be set to True
    :return: prepared set of entries
    """
    if isinstance(data_set, str):
        if quotes:
            return '"' + data_set + '"'
        else:
            return data_set
    elif isinstance(data_set, int):
        return data_set

    final_str = ""
    for i, entry in enumerate(data_set):
        if i < len(data_set) - 1:
            if isinstance(entry, int):
                final_str = final_str + str(entry) + ","
            elif not entry:
                final_str = final_str + "Null,"
            elif quotes:
                # SQLite only support quotes doubling, no escape is supported
                final_str = final_str + "'" + entry.replace("'", "''") + "',"
            else:
                final_str = final_str + entry + ","
        else:
            if isinstance(entry, int):
                final_str = final_str + str(entry)
            elif not entry:
                final_str = final_str + "Null"
            elif quotes:
                final_str = final_str + "'" + entry.replace("'", "''") + "'"
            else:
                final_str = final_str + entry
    return final_str


def do_query(qr_str, one_row=False, insert=False):
    """
    Execute a query string to the DB
    For select queries, one_row key value to select if to return only one row (if any) or a list of rows (if available)
    :param qr_str: query string that needs to be executed
    :param one_row: select to return only one row or all
    :param insert: if the call for insert query
    :return: return query results
    """

    if insert:
        return do_insert(qr_str)
    else:
        db_cur.execute(qr_str)
        if not one_row:
            return db_cur.fetchall()
        else:
            return db_cur.fetchone()


def do_insert(qr_str):
    """
    Excute insert query and commit
    :param qr_str: query string
    :return: last inserted row id
    """
    db_cur.execute(qr_str)
    db_con.commit()
    return db_cur.lastrowid


def insert_to_db(tbl_name, columns, value_tuple):
    """
    Execute an insert query, prepare table column name and values with quotes if necessary
    :param tbl_name: query table name
    :param columns: table columns to insert values to
    :param value_tuple: input data to a table row
    :return: last inserted row id
    """
    columns = prepare_string_with_quotes(columns)
    value_tuple = prepare_string_with_quotes(value_tuple, quotes=True)

    qr_str = 'insert into {0} ({1}) values ({2})'.format(tbl_name, columns, value_tuple)
    return do_insert(qr_str)


def insert_list_to_db_tbl(data_list, table_name, table_column, **kwargs):
    """
    Insert a data list into a DB table
    :param data_list:
    :param table_name:
    :param table_column:
    :param kwargs: key value pairs, keys for additional column name, values for row data
    :return:
    """
    for text in data_list:
        insert_to_db(table_name, [table_column] + list(kwargs.keys()), [text] + list(kwargs.values()))


def get_files_from_path(path, file_extension):
    """
    Get list of files from a path, with specific extention
    :param path: full directory path
    :param file_extension: to select files with that extension
    :return: list of files
    """
    files = [f for f in os.listdir(path) if isfile(join(path, f))]
    imp_files = list()
    for file in files:
        if fnmatch.fnmatch(file, '*' + file_extension):
            imp_files.append(file[:file.index('.')])

    return imp_files


def get_import_from_vrfs(vrf_id):
    """
    Get for a VRF X the import RTs and their corresponding VRF name and hostname
    :param vrf_id: current VRF X
    :return: a list of (filename, VRF)
    """
    mapped = []

    import_rts = do_query(""" select rt_import from import_rt where vrf_id = {0}""".format(vrf_id))

    for import_rt in import_rts:
        importers = do_query(
            """select site.name, vrf.name, vrf.rd, export_rt.rt_export
            from vrf
            join export_rt on vrf.vrf_id = export_rt.vrf_id
            join appliance on vrf.app_id = appliance.app_id
            join site on appliance.site_id = site.site_id
            where export_rt.rt_export = '{0}' and vrf.vrf_id != {1}; 
            """.format(import_rt[0], vrf_id))

        if importers:
            for imp in importers:
                mapped.append(imp)
        else:
            mapped.append(('N/A', 'N/A', 'N/A', import_rt[0]))
    return mapped


def get_exported_to(vrf_id):
    """
    Get for a VRF X the VRFs which are importing its exported RT
    :param vrf_id: VRF X db id
    :return: a list of (filename, VRF)
    """
    exported_to = []

    export_rts = do_query(""" select export_rt.rt_export from export_rt where vrf_id = {0}""".format(vrf_id))

    for export_rt in export_rts:
        who_import = do_query(
            """ select site.name, vrf.name, vrf.rd, export_rt.rt_export
                from vrf
                join appliance on vrf.app_id = appliance.app_id
                join site on appliance.site_id = site.site_id
                join import_rt on vrf.vrf_id = import_rt.vrf_id
                join export_rt on vrf.vrf_id = export_rt.vrf_id
                where import_rt.rt_import = "{0}" and vrf.vrf_id != {1};                 
            """.format(export_rt[0], vrf_id)
        )
        if who_import:
            for im in who_import:
                exported_to.append(im)

    return exported_to


def insert_vrf_to_db(sw_id, vrf_name, rd, imp_exp, description):
    """
    Insert complete VRF information into database
    :param sw_id: a foreign key app_id in vrf table
    :param vrf_name: vrf table, name column name
    :param rd: vrf table, rd column name
    :param imp_exp: list of lists(exports, imports)
    :param description: vrf table, description column name
    :return: nothing
    """
    vrf_db_id = insert_to_db('vrf', 'name, rd, description, app_id', (vrf_name, rd, description, sw_id))
    if len(imp_exp) == 2:
        # only one address family imports and exports: IPv4
        insert_list_to_db_tbl(imp_exp[0], 'export_rt', 'rt_export', vrf_id=vrf_db_id, add_fam_id=1)
        insert_list_to_db_tbl(imp_exp[1], 'import_rt', 'rt_import', vrf_id=vrf_db_id, add_fam_id=1)
    elif len(imp_exp) == 4:
        # for both IPv4 and IPv6
        if imp_exp[0] is not None:
            insert_list_to_db_tbl(imp_exp[0], 'export_rt', 'rt_export', vrf_id=vrf_db_id, add_fam_id=1)
        if imp_exp[1] is not None:
            insert_list_to_db_tbl(imp_exp[1], 'import_rt', 'rt_import', vrf_id=vrf_db_id, add_fam_id=1)
        if imp_exp[2] is not None:
            insert_list_to_db_tbl(imp_exp[2], 'export_rt', 'rt_export', vrf_id=vrf_db_id, add_fam_id=2)
        if imp_exp[3] is not None:
            insert_list_to_db_tbl(imp_exp[3], 'import_rt', 'rt_import', vrf_id=vrf_db_id, add_fam_id=2)


def parse_vrf(lines, index, vrf_name, vrf_rt_to_name, vrf_def=None, crs_asr=None):
    rd = ''
    description = ''
    index = index + 1
    maps = {}
    loop_break = False

    if not vrf_def and not crs_asr:
        # For IOS, old VRF configuration syntax.
        # Only for IPv4

        imports = []
        exports = []
        while index < len(lines):
            if not lines[index] and index + 1 > len(lines) and lines[index + 1]:
                index = index + 1
                continue

            if re.match(rgx_vrf_fwd, lines[index]) or \
                    re.match(rgx_ios_ip_vrf, lines[index]) or \
                    re.match(rgx_ios_vrf_def, lines[index]):
                break

            if re.match(rgx_ios_imp_rt, lines[index]):
                imports.append(re.match(rgx_ios_imp_rt, lines[index])['rt_import'])

            elif re.match(rgx_ios_exp_rt, lines[index]):
                export = re.match(rgx_ios_exp_rt, lines[index])['rt_export']
                exports.append(export)
                vrf_rt_to_name[export] = vrf_name

            elif re.match(rgx_rd, lines[index]):
                rd = re.match(rgx_rd, lines[index])['rd']

            elif re.match(rgx_description, lines[index]):
                description = re.match(rgx_description, lines[index])['description']

            elif re.match(rgx_import_map, lines[index]):
                maps['import'] = re.match(rgx_import_map, lines[index])['map']

            elif re.match(rgx_export_map, lines[index]):
                maps['export'] = re.match(rgx_export_map, lines[index])['map']

            elif re.match(r'\s+!\s*', lines[index]) or \
                    re.match(r'\s+\w+', lines[index]):
                if print_ignore_break:
                    print('Ignore line {0}: {1}'.format(index, lines[index]))
                pass
            else:
                if print_ignore_break:
                    print('break line: ' + lines[index])
                break
            index = index + 1
        return rd, [exports, imports], description, index

    elif vrf_def:
        # For IOS, new VRF configuration syntax
        # for both IPv4 and IPv6
        add_family = ''
        imports = []
        exports = []
        imports_v6 = []
        exports_v6 = []
        while index < len(lines):
            if not lines[index] and index + 1 > len(lines) and lines[index + 1]:
                index = index + 1
                continue
            if re.match(rgx_vrf_fwd, lines[index]) or \
                    re.match(rgx_ios_ip_vrf, lines[index]) or \
                    re.match(rgx_ios_vrf_def, lines[index]):
                break

            if re.match(rgx_ios_imp_rt, lines[index]):
                if add_family == 'IPv4':
                    imports.append(re.match(rgx_ios_imp_rt, lines[index])['rt_import'])
                elif add_family == 'IPv6':
                    imports_v6.append(re.match(rgx_ios_imp_rt, lines[index])['rt_import'])

            elif re.match(rgx_ios_exp_rt, lines[index]):
                if add_family == 'IPv4':
                    export = re.match(rgx_ios_exp_rt, lines[index])['rt_export']
                    exports.append(export)
                    vrf_rt_to_name[export] = vrf_name
                elif add_family == 'IPv6':
                    export = re.match(rgx_ios_exp_rt, lines[index])['rt_export']
                    exports_v6.append(export)
                    vrf_rt_to_name[export] = vrf_name
                else:
                    print('Something wrong in vrf definition in line number {0}'.format(index))

            elif re.match(rgx_rd, lines[index]):
                rd = re.match(rgx_rd, lines[index])['rd']

            elif re.match(rgx_add_family, lines[index]):
                add_family = 'IP' + re.match(rgx_add_family, lines[index])['add_family']

            elif re.match(rgx_description, lines[index]):
                description = re.match(rgx_description, lines[index])['description']

            elif re.match(rgx_import_map, lines[index]):
                if add_family == 'IPv4':
                    maps['importv4'] = re.match(rgx_import_map, lines[index])['map']
                elif add_family == 'IPv6':
                    maps['importv6'] = re.match(rgx_import_map, lines[index])['map']

            elif re.match(rgx_export_map, lines[index]):
                if add_family == 'IPv4':
                    maps['exportv4'] = re.match(rgx_export_map, lines[index])['map']
                elif add_family == 'IPv6':
                    maps['exportv6'] = re.match(rgx_export_map, lines[index])['map']

            elif re.match(r'\s+!\s*', lines[index]) or \
                    re.match(r'\s*exit-address-family\s*', lines[index]) or \
                    re.match(r'\s+\w+', lines[index]):
                if print_ignore_break:
                    print('Ignore line {0}: {1}'.format(index, lines[index]))

            else:
                if print_ignore_break:
                    print('break line: ' + lines[index], vrf_name)
                break
            index = index + 1

        return rd, [exports, imports, exports_v6, imports_v6], description, index

    else:
        # For IOS XR
        add_family = ''
        imports = []
        exports = []
        imports_v6 = []
        exports_v6 = []
        while index < len(lines):
            if not lines[index] and not index + 1 > len(lines):
                index = index + 1
                continue
            elif re.match(rgx_xr_vrf, lines[index]) or re.match(r'\S+', lines[index]):
                break

            elif re.match(rgx_add_family, lines[index]):
                add_family = 'IP' + re.match(rgx_add_family, lines[index])['add_family']
                index = index + 1
                while index < len(lines):
                    if re.match(rgx_xr_imp_rt, lines[index]):
                        while index + 1 < len(lines):
                            index = index + 1
                            rt = re.match(r'\s*(?P<rt>[0-9:]+)$', lines[index])
                            if rt:
                                if add_family == 'IPv4':
                                    imports.append(rt['rt'])
                                else:
                                    imports_v6.append(rt['rt'])
                            else:
                                loop_break = True
                                break

                    elif re.match(rgx_xr_exp_rt, lines[index]):

                        while index + 1 < len(lines):
                            index = index + 1
                            rt = re.match(r'\s*(?P<rt>[0-9:]+)$', lines[index])
                            if rt:
                                if add_family == 'IPv4':
                                    exports.append(rt['rt'])
                                else:
                                    exports_v6.append(rt['rt'])
                            else:
                                loop_break = True
                                break

                    elif re.match(r'\S+', lines[index]):
                        break

                    if not loop_break:
                        index = index + 1
                    else:
                        loop_break = False

            elif re.match(rgx_description, lines[index]):
                description = re.match(rgx_description, lines[index])['description']

            elif re.match(r'\s+!\s*', lines[index]) or \
                    re.match(r'\s+\w+', lines[index]):
                if print_ignore_break:
                    print('Ignore line {0}: {1}'.format(index, lines[index]))
                pass

            else:
                if print_ignore_break:
                    print('break line: ' + lines[index], vrf_name)
                break
            index = index + 1

        return [exports, imports, exports_v6, imports_v6], description, index


def get_subnet_id(ip_add):
    qr = do_query('select subnet_id from subnet where network_id="{0}"'.format(ip_add.network), True)
    if qr:
        return qr[0]
    else:
        return insert_to_db('subnet', 'network_id', [str(ip_add.network)])


def insert_ip_subnet(int_id, add_family_id, ipadd, add_type, pending=False):
    """
    Insert IP address into the corresponding DB table
        normal table (ip_address):          for known (complete) interface IP
        pending table (pend_ip_address): for unknown interface

    With the exception that some IP configuration is not directly assigned to local appliance interface
    inserted into normal table without int_id
        Ex. tunnel destination IP
    :param int_id: local appliance DB interface ID
    :param add_family_id: address family DB ID
    :param ipadd: IP address (string or ipaddress.interface object)
    :param add_type: "primary", "secondary", "vip"
    :param pending: choose which table to insert to
    :return: IP address DB id
    """
    if not pending:
        table_name = 'ip_address'
    else:
        table_name = 'pend_ip_address'
    if int_id:
        if isinstance(ipadd, str):
            ipadd = ipaddress.ip_interface(ipadd)

        sub_id = get_subnet_id(ipadd)
        ip_qr = do_query(
            'select ip_id from ip_address where address="{0}" and int_id={1} and subnet_id={2};'.format(
                ipadd.with_prefixlen, int_id, sub_id), one_row=True)
        if ip_qr:
            return ip_qr[0]
        else:
            return insert_to_db(table_name, 'add_fam_id, address, subnet_id, address_type, int_id',
                                [add_family_id, ipadd.with_prefixlen, sub_id, add_type, int_id])
    else:
        # only normal ip_address table as the IP is not linked to any pending interfaces
        ip_qr = do_query(
            'select ip_id from ip_address '
            'where address="{0}" and (int_id is null or int_id="");'.format(ipadd), one_row=True)
        if ip_qr:
            return ip_qr[0]
        else:
            return insert_to_db('ip_address', 'add_fam_id, address, address_type',
                                [add_family_id, ipadd, add_type])


def get_most_specific_subnet(ip_add, sw_id):
    most_specific = ()
    subnets = do_query(
        """select subnet.network_id, subnet.subnet_id
                from ip_address
                join subnet on ip_address.subnet_id = subnet.subnet_id
                join interface on ip_address.int_id = interface.int_id
                where interface.app_id = {0};""".format(sw_id)
    )
    if not subnets:
        return None

    for subnet in subnets:
        net = ipaddress.ip_network(subnet[0])
        if ip_add in net:
            if most_specific:
                if ipaddress.ip_network(most_specific[0]).prefixlen < net.prefixlen:
                    most_specific = subnet
            else:
                most_specific = subnet
    return most_specific[1]


def get_vlan_id(vlan, sw_id):
    vlan_db_id = ()
    if isinstance(vlan, str):
        try:
            vlan = int(vlan)
            vlan_db_id = do_query(
                'select vlan_id from vlan where vlan_no={0} and app_id={1};'.format(vlan, sw_id), True)
            if vlan == 1 and not vlan_db_id:
                vlan_db_id = [insert_to_db('vlan', 'vlan_no, name, app_id, exist', [1, 'Default', sw_id, 1])]
        except ValueError:
            # given a vlan name, get its DB id
            vlan_db_id = do_query(
                'select vlan_id from vlan where name={0} and app_id={1};'.format(vlan, sw_id), True)
    elif isinstance(vlan, int):
        vlan_db_id = do_query(
            'select vlan_id from vlan where vlan_no={0} and app_id={1};'.format(vlan, sw_id), True)
        if vlan == 1 and not vlan_db_id:
            vlan_db_id = [insert_to_db('vlan', 'vlan_no, name, app_id, exist', [1, 'Default', sw_id, 1])]

    if vlan_db_id:
        return vlan, vlan_db_id[0]
    else:
        return vlan, None


def find_vrf_interfaces(sw_id, vrf_id, include_default_vrf=False):
    """
    Find VRF related interfaces, groups all assigned IP address to each interface
    :param sw_id: appliance DB ID
    :param vrf_id: vrf DB ID
    :param include_default_vrf: if to get default vrf interfaces or not, default is not
    :return: list of tuples (vrf_name, vrf_export_rt(s), interface_name_type, ip_address(es), interface_description)
    """
    if include_default_vrf:
        table_join = 'left outer join'
    else:
        table_join = 'join'

    interfaces = do_query(
        """ select vrf_name, group_concat(export_rt), interface, ip, description, sw_id, int_type, 
            int_number, id, status 
            from
            (
            select vrf.name as vrf_name, export_rt.rt_export as export_rt,
                case
                    when interface.type='vlan' then upper(interface.type)||interface.number
                    else (interface.type||interface.number)	
                end as interface, 
            group_concat(ip_address.address, ',') as ip, interface.description as description, 
            appliance.app_id as sw_id, interface.type as int_type, interface.number as int_number, 
            interface.int_id as id, interface.status as status
            from interface
            join vrf on interface.vrf_id = vrf.vrf_id
            join appliance on appliance.app_id = interface.app_id
            join ip_address on ip_address.int_id = interface.int_id
            join subnet on subnet.subnet_id = ip_address.subnet_id
            {2} export_rt on export_rt.vrf_id = interface.vrf_id
            where vrf.vrf_id={0} and appliance.app_id={1}
            group by vrf.name, interface, export_rt.rt_export, interface.description
            ) as vrfs

            group by vrf_name, interface, description;
        """.format(vrf_id, sw_id, table_join))
    list_interfaces = []
    if interfaces:
        for interface in interfaces:
            # convert a tuple to list
            int_list = [i for i in interface]
            if ',' in int_list[3]:
                intt = int_list[3].split(',')
                int_list[3] = intt[0] + ', ' + intt[1][:intt[1].index('/')]

            if int_list[6] == 'vlan':
                qr = do_query(
                    'select name from vlan '
                    'where vlan_no = {0} and app_id = {1} and '
                    'exist = 1;'.format(int_list[7], sw_id), True)

                if qr and qr[0]:
                    int_list[2] = int_list[2] + ' ({0})'.format(qr[0].strip())

            if '.' in int_list[2]:
                int_qr = do_query('select vlan.vlan_no from vlan '
                                  'join int_vlan on vlan.vlan_id=int_vlan.vlan_id '
                                  'join interface on int_vlan.int_id=interface.int_id '
                                  'where interface.int_id={0} and vlan.exist=1;'.format(int_list[8]), True)
                if int_qr:
                    int_list[2] = int_list[2] + ' ({0})'.format(int_qr[0])

            if int_list[9] == 'shutdown':
                int_list[2] = int_list[2] + ' (Shutdown)'

            if int_list[6] == 'Port-Channel':
                # qr = do_query('select group_concat(interface.type||interface.number ) '
                #               'from interface where interface.member_of={0}'.format(int_list[8]), True)
                # int_list[2] = int_list[2] + ' ({0})'.format(qr[0])
                qr = do_query('select group_concat(substr(interface.type, 1, 3)||interface.number) '
                              'from interface where interface.member_of={0}'.format(int_list[8]), True)
                int_list[2] = int_list[2] + ' ({0})'.format(qr[0])
            list_interfaces.append(int_list)

        list_interfaces.sort(key=lambda item: sort_interface_num_key(item[7]))
    return list_interfaces


def find_static_routes_summary(sw_id, vrf_id, include_null_values=True):
    if include_null_values:
        table_join = 'left outer join'
    else:
        table_join = 'join'

    qr = do_query(
        """
        select ip_address.address as next_hop, substr(interface.type, 1, 3)||interface.number as interface, 
        count(*) as c, static_route.to_vrf_id

        from static_route
        join subnet on static_route.subnet_id = subnet.subnet_id
        {0} ip_address on static_route.next_hop_ip = ip_address.ip_id
        {0} interface on static_route.next_hop_int = interface.int_id
        where static_route.app_id = {1} and static_route.vrf_id= {2}
        group by ip_address.address, interface
        order by c DESC;
        """.format(table_join, sw_id, vrf_id))
    if qr:
        next_hops = []
        to_vrf = ''

        for next_hop in qr:
            if next_hop[3]:
                # next hop in different vrf
                to_vrf_qr = do_query('select vrf.name from vrf where vrf.vrf_id={0};'.format(next_hop[3]), True)
                if to_vrf_qr and to_vrf_qr[0]:
                    to_vrf = ' ->({0})'.format(to_vrf_qr[0])

            if next_hop[0] and next_hop[1]:
                next_hop_str = next_hop[0][:next_hop[0].index('/')] + ' ({0})'.format(next_hop[1]) + to_vrf

            elif next_hop[0]:
                next_hop_str = next_hop[0][:next_hop[0].index('/')] + to_vrf

            else:
                next_hop_str = next_hop[1] + to_vrf

            next_hops.append((next_hop_str, next_hop[2]))
        return next_hops
    return ()


def parse_interface(sw_id, lines, index, int_type_name='', int_number=''):
    """
    port-channel, loopback, and normal interface ports

    :param sw_id:
    :param lines:
    :param index:
    :param int_type_name:
    :param int_number:
    :return:
    """
    vrf_name = 'Default'
    if len(do_query('select * from vrf where name="{0}" and app_id={1}'.format(vrf_name, sw_id))) == 0:
        # make sure default VRF are inserted into the DB
        insert_to_db('vrf', 'name, app_id', [vrf_name, sw_id])

    sub_int_num = ''
    encap_vlan_id = ''

    if not int_type_name and not int_number:
        match = re.match(rgx_interface, lines[index])
        int_type_name = match['type']
        int_number = match['number']
        sub_int_num = match['sub'] if match['sub'] else ""

    int_type = ''  # l2 or l3
    mode = 'virtual' if int_type_name.lower() == 'loopback' else ''  # access, trunk or virtual (vlan, tunnel)
    status = 'up'  # up or shutdown

    description = ''
    ip_add = []
    allowed_vlans_ids = []
    native_vlan_id = ''

    access_vlans_ids = []
    group_id = 0
    group_mode = ''

    index = index + 1
    while lines[index] != '!' and index < len(lines):
        if re.match(rgx_description, lines[index]):
            description = re.match(rgx_description, lines[index])['description']

        elif re.match(r'\s+encapsulation\s+dot1q\s+(?P<id>\d+)', lines[index]):
            match_vlan = re.match(r'\s+encapsulation\s+dot1q\s+(?P<id>\d+)', lines[index])
            encap_vlan_id = get_vlan_id(match_vlan['id'].strip(), sw_id)
            if encap_vlan_id:
                # get_vlan_id returns (vlan_str, vlan_db_id)
                encap_vlan_id = encap_vlan_id[1]
            if not encap_vlan_id:
                encap_vlan_id = insert_to_db('vlan', 'vlan_no, name, app_id, exist', [match_vlan['id'], "", sw_id, 1])

        elif check_interface_vrf_forwarding(lines[index]):
            if not int_type:
                int_type = 'l3'
            elif int_type != 'l3':
                print('Interface type may not be accurate for the interface "{0}" on line: {1}'.format(
                    int_type_name + int_number, index))

            vrf_name = match_interface_vrf_forwarding(lines[index])

        elif check_if_ip_address(lines[index]):
            if not int_type:
                int_type = 'l3'
            elif int_type != 'l3':
                print('Interface type may not be accurate for the interface "{0}" on line: {1}'.format(
                    int_type_name + int_number, index))

            ip_add.append(match_ip_address(lines[index])[1])

        elif re.match(r'\s+switchport\s*.*', lines[index]):
            if not int_type:
                int_type = 'l2'
            elif int_type != 'l2':
                print('Interface type may not be accurate for the interface "{0}" on line: {1}'.format(
                    int_type_name + int_number, index))

            index, mode, access_vlans_ids, allowed_vlans_ids, native_vlan_id = match_interface_switchport_statements(
                sw_id, lines, index)

        elif re.match(r'\s*channel-group\s*(?P<group_id>\d+)(\s*mode\s*(?P<mode>.+?))?\s*', lines[index]):
            match = re.match(r'\s*channel-group\s*(?P<group_id>\d+)(\s*mode\s*(?P<mode>.+?))?\s*', lines[index])
            group_id = match['group_id']
            group_mode = match['mode'] if match['mode'] else ''

        elif re.match(r'\s*shutdown\s*$', lines[index]):
            status = 'shutdown'

        elif re.match(r'\s+\w+', lines[index]):
            # ignore unneeded lines under interface
            if print_ignore_break:
                print('Ignore line {0}: {1}'.format(index, lines[index]))
            pass

        elif is_interface_section_end(lines[index]):
            if print_ignore_break:
                print('Break line: ' + lines[index])
            break
        index = index + 1

    int_vlan_table = 'int_vlan'
    interface_table = 'interface'
    pending = False
    pending_reason = 'port-channel'

    if int_type_name == 'Port-Channel':
        int_id = insert_to_db('interface',
                              'type, number, description, mode, app_id, status',
                              [int_type_name, int_number, description, mode, sw_id, status])
    else:
        if group_id:
            # if member of port-channel
            group_int_id = do_query(
                'select int_id from interface where type="Port-Channel" and number={0} and app_id={1}'.format(
                    group_id, sw_id), True)

            if group_int_id:
                int_id = insert_to_db(interface_table,
                                      'type, number, description, mode, member_of, app_id, status',
                                      [int_type_name, int_number, description, mode, group_int_id[0], sw_id, status])
            else:
                pending = True
                int_vlan_table = 'pend_int_vlan'  # pending because VLAN is configured for a pending interface
                interface_table = 'pending_interface'
                int_id = insert_to_db(interface_table,
                                      'type, number, description, mode, group_id, app_id, pend_reason, status',
                                      [int_type_name, int_number, description, mode,
                                       group_id, sw_id, pending_reason, status])
        else:
            # is not a member of any port-channel
            int_id = insert_to_db(interface_table,
                                  'type, number, description, mode, app_id, status',
                                  [int_type_name, int_number, description, mode, sw_id, status])

    if mode == 'trunk':
        if not native_vlan_id:
            native_vlan_id = get_vlan_id('1', sw_id)

        for vlan in allowed_vlans_ids:
            if not vlan[1]:
                # configured with allowed vlan but the vlan id doesn't exist
                vlan_id = insert_to_db('vlan', 'vlan_no, app_id, exist',
                                       [int(vlan[0]) if isinstance(vlan[0], str) else vlan[0], sw_id, 0])

                insert_to_db(int_vlan_table, 'int_id, vlan_id, vlan_mode', [int_id, vlan_id, 'tagged'])
            else:
                insert_to_db(int_vlan_table, 'int_id, vlan_id, vlan_mode', [int_id, vlan[1], 'tagged'])

        if not native_vlan_id[1]:
            # configured with native vlan but the vlan id doesn't exist
            vlan_id = insert_to_db('vlan', 'vlan_no, app_id, exist',
                                   [int(native_vlan_id[0]) if isinstance(native_vlan_id[0], str) else native_vlan_id[0],
                                    sw_id, 0])

            insert_to_db(int_vlan_table, 'int_id, vlan_id, vlan_mode', [int_id, vlan_id, 'untagged'])
        else:
            insert_to_db(int_vlan_table, 'int_id, vlan_id, vlan_mode', [int_id, native_vlan_id[1], 'untagged'])

    elif mode == 'access':
        # mode access
        for vlan in access_vlans_ids:
            if vlan[1] or int_vlan_table == 'pending_int_vlan':
                insert_to_db(int_vlan_table, 'int_id, vlan_id, vlan_mode', [int_id, vlan[0][1], vlan[1]])
            else:
                # vlan id is not available in DB
                vlan_id = insert_to_db('vlan', 'vlan_no, app_id, exist',
                                       [int(vlan[0]) if isinstance(vlan[0], str) else vlan[0], sw_id, 0])

                insert_to_db(int_vlan_table, 'int_id, vlan_id, vlan_mode', [int_id, vlan_id, vlan[1]])

    else:
        # for L3 interfaces
        qr = do_query('select vrf_id from vrf where name="{0}" and app_id={1}'.format(vrf_name, sw_id), True)
        if qr:
            if pending:
                # pending_interface table doesn't have vrf_id field, adding that reason to update the other interface
                # table with it later
                pending_reason = pending_reason + ',vrf_id'
                do_query(
                    'update {0} set vrf_name="{1}", pend_reason="{2}" where int_id={3}'.format(
                        interface_table, vrf_name, pending_reason, int_id))

            if not pending:
                vrf_id = qr[0]
                # update the interface record wih vrf_id
                do_query('update {0} set vrf_id={1} where int_id={2}'.format(interface_table, vrf_id, int_id))
        else:
            # update pending reason with vrf_id is unknown
            if pending:
                pending_reason = pending_reason + ',vrf_id'
                do_query(
                    'update {0} set vrf_name={1}, pend_reason={2} where int_id={3}'.format(
                        interface_table, vrf_name, pending_reason, int_id))

            else:
                """
                1) read the record from interface table
                3) update it with pending reason and save to pend_interface table
                2) delete int_id record in normal interface table                    
                """
                pending = True
                pending_reason = 'vrf_id'
                do_query(
                    'insert into pending_interface '
                    '(type, number, description, mode, member_of, app_id, vrf_name, pend_reason, status) '
                    'select '
                    'type, number, description, mode, member_of, app_id, '
                    '"{0}" as vrf_name, "{1}" as pend_reason, status '
                    'from interface where int_id={2}'.format(vrf_name, pending_reason, int_id), insert=True)

                do_query('delete from interface where int_id={0}'.format(int_id))

        if not pending:
            for ip_entry in ip_add:
                insert_ip_subnet(int_id, ip_entry[0], ip_entry[1], ip_entry[2])

            if sub_int_num:
                insert_to_db(int_vlan_table, 'int_id, vlan_id, vlan_mode', [int_id, encap_vlan_id, 'untagged-access'])

        else:
            for ip_entry in ip_add:
                insert_ip_subnet(int_id, ip_entry[0], ip_entry[1], ip_entry[2], pending=True)

            if sub_int_num:
                int_vlan_table = 'pending_int_vlan'
                insert_to_db(int_vlan_table, 'int_id, vlan_id, vlan_mode', [int_id, encap_vlan_id, 'untagged-access'])

    return index


def match_interface_switchport_statements(sw_id, lines, index):
    access_vlans_ids = []
    allowed_vlans_ids = []
    mode = ''
    native_vlan_id = ''

    while index < len(lines):

        if re.match(r'\s*switchport\s*mode\s*(?P<mode>.+?)', lines[index]):
            mode = re.match(r'\s*switchport\s*mode\s*(?P<mode>.+)', lines[index])['mode'].strip()

        elif re.match(r'\s*switchport\s+access\s+vlan\s+(?P<access>.+)', lines[index]):
            access_vlan = get_vlan_id(
                re.match(r'\s*switchport\s+access\s+vlan\s+(?P<access>.+)', lines[index])['access'].strip(), sw_id)
            access_vlans_ids.append([access_vlan, 'untagged-access'])

        elif re.match(r'\s*switchport\s+voice\s+vlan\s+(?P<voice>.+)', lines[index]):
            voice_vlan = get_vlan_id(
                re.match(r'\s*switchport\s+voice\s+vlan\s+(?P<voice>.+)', lines[index])['voice'].strip(), sw_id)
            access_vlans_ids.append([voice_vlan, 'untagged-voice'])

        elif re.match(r'\s*switchport\s+trunk\s+allowed\s+vlan\s+(?P<ids>\d+(-\d+)?(,\d+(-\d+)?)*)', lines[index]):
            match = re.match(
                r'\s*switchport\s+trunk\s+allowed\s+vlan\s+(?P<ids>\d+(-\d+)?(,\d+(-\d+)?)*)', lines[index])
            allowed_vlans_ids.extend(get_vlan_id(int(vlan), sw_id) for vlan in get_vlan_numbers(match['ids']))

        elif re.match(r'\s*switchport\s+trunk\s+allowed\s+vlan\s+add\s+(?P<ids>\d+(-\d+)?(,\d+(-\d+)?)*)', lines[index]):
            match = re.match(
                r'\s*switchport\s+trunk\s+allowed\s+vlan\s+add\s+(?P<ids>\d+(-\d+)?(,\d+(-\d+)?)*)', lines[index])
            allowed_vlans_ids.extend(get_vlan_id(int(vlan), sw_id) for vlan in get_vlan_numbers(match['ids']))

        elif re.match(r'\s*switchport\s+trunk\s+native\s+vlan\s+(?P<native>\d+)', lines[index]):
            match = re.match(r'\s*switchport\s+trunk\s+native\s+vlan\s+(?P<native>\d+)', lines[index])
            native_vlan_id = get_vlan_id(match['native'], sw_id)

        if re.match(r'\s+switchport\s*.*', lines[index + 1]):
            index = index + 1
        else:
            break

    return index, mode, access_vlans_ids, allowed_vlans_ids, native_vlan_id


def check_interface_vrf_forwarding(line):
    """
    Returns IOS version if the passed line assigns the current interface to a VRF
    Different IOS has different syntax

    :param line: configuration line passed from an interface parsing function
    :return: ios_ver
    """

    if re.match(rgx_vrf_fwd, line):
        ios_ver = 'ios'

    elif re.match(rgx_xr_vrf, line):
        ios_ver = 'xr'

    elif re.match(rgx_nx_vrf_fwd, line):
        ios_ver = 'nx'

    else:
        return False

    return ios_ver


def match_interface_vrf_forwarding(line):
    ver = check_interface_vrf_forwarding(line)
    if ver == 'ios':
        vrf_name = re.match(rgx_vrf_fwd, line)['vrf_name']

    elif ver == 'xr':
        vrf_name = re.match(rgx_xr_vrf, line)['vrf_name']

    elif ver == 'nx':
        vrf_name = re.match(rgx_nx_vrf_fwd, line)['vrf_name']

    else:
        return False

    return vrf_name


def is_interface_section_end(line):
    if re.match(r'!\s*', line) or re.match(r'\s*$', line) or re.match(r'\S', line):
        return True
    else:
        return False


def check_if_ip_address(line):

    if re.match(rgx_ipv6_add, line):
        # IPv6 address
        return 6

    elif re.match(rgx_nx_ip_add, line):
        return 4, 'nx'

    elif re.match(rgx_ip_add, line):
        # IPv4 address
        return 4

    else:
        return False


def match_ip_address(line):
    ver = check_if_ip_address(line)

    if not isinstance(ver, tuple) and ver == 4:
        # IPv4 address
        match = re.match(rgx_ip_add, line)
        ip = match['ip']
        subnet = match['subnet']

        if match.group('sec'):
            return subnet, [add_family_v4_id, ipaddress.ip_interface(ip + '/' + subnet), 'secondary']
        else:
            return subnet, [add_family_v4_id, ipaddress.ip_interface(ip + '/' + subnet), 'primary']

    elif isinstance(ver, tuple) and ver[0] == 4 and ver[1] == 'nx':
        match = re.match(rgx_nx_ip_add, line)
        ip = match['ip']
        subnet = match['subnet']

        if match.group('sec'):
            return subnet, [add_family_v4_id, ipaddress.ip_interface(ip + '/' + subnet), 'secondary']
        else:
            return subnet, [add_family_v4_id, ipaddress.ip_interface(ip + '/' + subnet), 'primary']

    elif ver == 6:
        # IPv6 address
        match = re.match(rgx_ipv6_add, line)
        ip = match['ip_add']
        subnet = match['subnet'] if match['subnet'] else 64
        ip_type = match['type'] if match['type'] else ""

        if ip_type:
            return subnet, [add_family_v6_id, ipaddress.ip_interface(ip), ip_type]
        else:
            return subnet, [add_family_v6_id, ipaddress.ip_interface(ip), 'global']

    else:
        return False


def parse_interface_vlan(sw_id, lines, index):
    ip_add = []
    vrf_name = 'Default'
    mode = 'virtual'
    status = 'up'  # up or shutdown

    subnet = ''
    description = ''

    vlan_id = re.match(r'\s*[iI]nterface\s+[vV]lan(?P<number>.+)', lines[index])['number']

    index = index + 1
    while index < len(lines) and lines[index] != '!':

        if not lines[index].strip():
            break

        if re.match(rgx_description, lines[index]):
            description = re.match(rgx_description, lines[index])['description']

        elif check_interface_vrf_forwarding(lines[index]):
            vrf_name = match_interface_vrf_forwarding(lines[index])

        elif check_if_ip_address(lines[index]):
            subnet, ips = match_ip_address(lines[index])
            ip_add.append(ips)

        elif re.match(r'\s*standby\s*\d+\s+ip\s+(?P<vip>\d+\.\d+\.\d+\.\d+)', lines[index]):
            vip = re.match(r'\s*standby\s*\d+\s+ip\s+(?P<vip>\d+\.\d+\.\d+\.\d+)', lines[index])['vip']
            ip_add.append([add_family_v4_id, ipaddress.ip_interface(vip + '/' + subnet if subnet else '32'), 'vip'])

        elif re.match(r'\s+hsrp\s+(?P<group>\d+)', lines[index]):
            while index + 1 < len(lines):
                index = index + 1
                match = re.match(rgx_ip_add, lines[index])
                if match:
                    vip = match['ip']
                    ip_add.append([add_family_v4_id,
                                   ipaddress.ip_interface(vip + '/' + subnet if subnet else '32'), 'vip'])
                    break
                elif re.match(r'\s+\S+', lines[index]):
                    continue
                elif is_interface_section_end(lines[index]):
                    break

        elif re.match(r'\s*shutdown\s*$', lines[index]):
            status = 'shutdown'

        elif re.match(r'\s+\w+', lines[index]):
            # ignore unneeded lines under interface
            if print_ignore_break:
                print('Ignore line {0} {2}: {1}'.format(index, lines[index], sw_id))

        elif is_interface_section_end(lines[index]):
            if print_ignore_break:
                print('break line: ' + lines[index])
            break

        index = index + 1

    pending = False
    pending_reason = 'vrf_id'

    qr = do_query('select vrf_id from vrf where name="{0}" and app_id={1}'.format(vrf_name, sw_id), one_row=True)
    if vrf_name == 'Default' and not qr:
        vrf_id = insert_to_db('vrf', 'name, app_id', [vrf_name, sw_id])
        int_id = insert_to_db('interface', 'type, number, description, mode, app_id, vrf_id, status',
                              ['vlan', vlan_id, description, mode, sw_id, vrf_id, status])

    else:
        if qr:
            vrf_id = qr[0]
            int_id = insert_to_db('interface',
                                  'type, number, description, mode, app_id, vrf_id, status',
                                  ['vlan', vlan_id, description, mode, sw_id, vrf_id, status])
        else:
            int_id = insert_to_db('pending_interface',
                                  'type, number, description, mode, app_id, vrf_name, pend_reason, status',
                                  ['vlan', vlan_id, description, mode, sw_id, vrf_name,
                                   pending_reason, status])
            pending = True

    if not pending:
        for ip_entry in ip_add:
            insert_ip_subnet(int_id, ip_entry[0], ip_entry[1], ip_entry[2])
    else:
        for ip_entry in ip_add:
            insert_ip_subnet(int_id, ip_entry[0], ip_entry[1], ip_entry[2], pending=True)

    return index


def parse_interface_tunnel(sw_id, lines, index):
    tun_num = re.match(rgx_tunnel_interface, lines[index])['number']
    tun_source = ()
    tun_dest_ip = ''
    ip_add = []
    description = ''
    vrf_name = 'Default'
    mode = 'virtual'
    status = 'up'  # up or shutdown

    index = index + 1
    while lines[index] != '!' and index < len(lines):
        if re.match(rgx_description, lines[index]):
            description = re.match(rgx_description, lines[index])['description']

        elif check_interface_vrf_forwarding(lines[index]):
            vrf_name = match_interface_vrf_forwarding(lines[index])

        elif check_if_ip_address(lines[index]):
            ip_add.append(match_ip_address(lines[index])[1])

        elif re.match(r'\s*tunnel\s+source\s+(?P<tun_src>\d+\.\d+\.\d+\.\d+)', lines[index]):
            match = re.match(r'\s*tunnel\s+source\s+(?P<tun_src>\d+\.\d+\.\d+\.\d+)', lines[index])
            tun_source = ('source_ip', match['tun_src'])

        elif re.match(r'\s*tunnel\s+source\s+(?P<tun_src>(?P<type>[^0-9]+)(?P<number>.+))', lines[index]):
            match = re.match(r'\s*tunnel\s+source\s+(?P<tun_src>(?P<type>[^0-9]+)(?P<number>.+))', lines[index])
            tun_source = ('source_int', match['tun_src'])

        elif re.match(r'\s*tunnel\s+destination\s+(?P<tun_dest>\d+\.\d+\.\d+\.\d+)', lines[index]):
            match = re.match(r'\s*tunnel\s+destination\s+(?P<tun_dest>\d+\.\d+\.\d+\.\d+)', lines[index])
            tun_dest_ip = match['tun_dest']

        elif re.match(r'\s*shutdown\s*$', lines[index]):
            status = 'shutdown'

        elif re.match(r'\s+\w+', lines[index]):
            # ignore unneeded lines under interface
            if print_ignore_break:
                print('Ignore line {0}: {1}'.format(index, lines[index]))
            pass

        elif is_interface_section_end(lines[index]):
            if print_ignore_break:
                print('break line: ' + lines[index])
            break

        index = index + 1

    tun_id = insert_to_db('pending_tunnel_int', '{0}, dest_ip'.format(tun_source[0]), [tun_source[1], tun_dest_ip])
    int_id = insert_to_db('pending_interface',
                          'type, number, description, mode, app_id, vrf_name, tunnel_id, status',
                          ['tunnel', tun_num, description, mode, sw_id, vrf_name, tun_id, status])

    for ip_entry in ip_add:
        insert_ip_subnet(int_id, ip_entry[0], ip_entry[1], ip_entry[2], pending=True)

    insert_ip_subnet(None, add_family_v4_id, tun_dest_ip + '/32', 'tunnel-destination')

    return index


def parse_all_interface_type(sw_id, ios_ver, lines, index):

    if re.match(r'\s*[iI]nterface\s+[vV]lan(?P<number>.+)', lines[index]):
        index = parse_interface_vlan(sw_id, lines, index)

    elif re.match(rgx_tunnel_interface, lines[index]):
        index = parse_interface_tunnel(sw_id, lines, index)

    elif re.match(r'\s*[iI]nterface\s+[pP]ort-channel(?P<number>.+)', lines[index]):
        number = re.match(r'\s*[iI]nterface\s+[pP]ort-channel(?P<number>.+)', lines[index])['number']
        index = parse_interface(sw_id, lines, index + 1, int_type_name='Port-Channel', int_number=number)

    else:
        # match = re.match(rgx_interface, lines[index])
        index = parse_interface(sw_id, lines, index)

    return index + 1


def get_vlan_numbers(ids_str):
    vlan_ids = []
    if ',' in ids_str:
        vlans = ids_str.split(',')
        for vlan in vlans:
            if '-' in vlan:
                vv = vlan.split('-')
                for i in range(int(vv[0]), int(vv[1]) + 1):
                    vlan_ids.append(i)
            else:
                vlan_ids.append(int(vlan))
    elif '-' in ids_str:
        for i in range(int(ids_str[:ids_str.index('-')]), int(ids_str[ids_str.index('-') + 1:]) + 1):
            vlan_ids.append(i)
    else:
        vlan_ids.append(int(ids_str))

    return vlan_ids


def parse_vlan(lines, index, sw_id):
    match = re.match(r'\s*vlan\s+(?P<ids>\d+(-\d+)?(,\d+(-\d+)?)*)', lines[index])['ids']
    vlan_ids = get_vlan_numbers(match)

    name = ''
    index = index + 1
    while index < len(lines):
        if re.match(r'\s*name\s*(.+)', lines[index]):
            name = re.match(r'\s*name\s*(?P<name>.+)', lines[index])['name'].strip()
        elif re.match(r'\s+\w+', lines[index]):
            # ignore unneeded lines under interface
            if print_ignore_break:
                print('Ignore line {0}: {1}'.format(index, lines[index]))
            pass
        elif re.match(r'\s*!', lines[index]) or re.match(r'\S+', lines[index]):
            if print_ignore_break:
                print('break line: ' + lines[index])
            break
        index = index + 1

    for vlan_id in vlan_ids:
        if not do_query('select vlan_id from vlan where vlan_no={0} and app_id={1};'.format(vlan_id, sw_id), True):
            insert_to_db('vlan', 'vlan_no, name, app_id, exist', [vlan_id, name, sw_id, 1])

    return index


def parse_static_route(lines, index, sw_id):
    match = re.match(rgx_ios_static_route, lines[index])

    vrf_name = match['vrf']
    if not vrf_name:
        vrf_name = 'Default'
    ip_next_hop = match['ip_next_hop']
    int_type = match['type'].lower() if match['type'] else match['type']
    int_number = match['number']
    if match['sub_int']:
        int_number = int_number + match['sub_int']
    ad = match['ad']
    if not ad:
        ad = 1
    net = match['net']
    mask = match['mask']
    name = match['name']
    if not match['add_fam']:
        # IPv4
        add_fam = 1
    else:
        # IPv6
        add_fam = 2
    sub_id = get_subnet_id(ipaddress.ip_interface(net + '/' + mask))
    insert_to_db(
        'pend_static_route',
        'next_hop_ip, next_hop_int_type, next_hop_int_number, vrf_name, '
        'subnet_id, ad_distance, name, add_fam_id, app_id',
        [ip_next_hop, int_type, int_number, vrf_name, sub_id, ad, name, add_fam, sw_id])

    return index + 1


def parse_xr_static_route_entry(sw_id, match, vrf_name, add_fam='IPv4'):
    net = match['net']
    mask = match['mask']
    sub_id = get_subnet_id(ipaddress.ip_interface(net + '/' + mask))

    to_vrf = match['vrf'] if match['vrf'] else None

    int_type = match['type'] if match['type'] else ""
    int_number = match['number'] if match['number'] else ""
    if match['sub_int']:
        int_number = int_number + match['sub_int']

    ip_next_hop = match['ip_next_hop'] if match['ip_next_hop'] else ""
    ad = match['ad'] if not match['ad'] else ""
    description = match['description'] if match['description'] else ""

    insert_to_db(
        'pend_static_route',
        'next_hop_ip, next_hop_int_type, next_hop_int_number, vrf_name, '
        'subnet_id, ad_distance, name, add_fam_id, app_id, to_vrf_name',
        [ip_next_hop, int_type, int_number, vrf_name, sub_id, ad, description, add_fam, sw_id, to_vrf])


def parse_xr_vrf_address_family_section(sw_id, lines, index, vrf='Default'):

    add_fam = 'IP' + re.match(rgx_add_family, lines[index])['add_family']

    index = index + 1
    while index < len(lines):
        if re.match(rgx_xr_static_route_entry, lines[index]):
            match = re.match(rgx_xr_static_route_entry, lines[index])
            parse_xr_static_route_entry(sw_id, match, vrf, add_fam)

        elif re.match(r'\s*!', lines[index]):
            break
        else:
            print("1 Unkown static route line, index:{0}, line: {1}".format(index, lines[index]))

        index = index + 1

    return index - 1  # as index will be increased after return, if not -1 one line will be passed without processing


def parse_xr_static_route(lines, index, sw_id):

    rgx_vrf = r'\svrf\s+(?P<vrf>.+)'

    index = index + 1

    while index < len(lines):
        if re.match(rgx_add_family, lines[index]):
            # Under default VRF
            index = parse_xr_vrf_address_family_section(sw_id, lines, index)

        elif re.match(rgx_vrf, lines[index]):
            vrf = re.match(rgx_vrf, lines[index])['vrf']
            index = index + 1

            while index < len(lines):

                if re.match(rgx_add_family, lines[index]):
                    # Under a named VRF
                    index = parse_xr_vrf_address_family_section(sw_id, lines, index, vrf)

                elif not re.match(r'\s+!', lines[index]):
                    print('3 Unkown static route line, index:{0}, line: "{1}"'.format(index, lines[index]))

                elif re.match(r'\s!', lines[index]):
                    break

                index = index + 1

        elif re.match(r'\S+', lines[index]) or re.match(rgx_vrf, lines[index]):
            return index

        index = index + 1


def parse_xr_router_bgp(lines, index, sw_id):
    index = index + 1

    while index < len(lines):
        match = re.match(r'\svrf\s(?P<vrf_name>.+)', lines[index])
        if match:
            vrf_name = match['vrf_name']
            while index < len(lines):
                index = index + 1
                rd_match = re.match(r'\s\srd\s(?P<rd>\d+(:\d+)?)', lines[index])
                if rd_match:
                    # update the VRF RD into the DB
                    vrf_id = do_query(
                        'select vrf.vrf_id, vrf.rd, appliance.app_id '
                        'from vrf '
                        'join appliance on vrf.app_id = appliance.app_id '
                        'where appliance.app_id = {0} and vrf.name="{1}";'.format(sw_id, vrf_name), one_row=True)
                    if vrf_id:
                        insert_to_db('vrf', 'rd', [rd_match['rd'].strip()])
                    else:
                        # add this VRF, print notification
                        insert_to_db('vrf', 'name, rd, app_id', [vrf_name, rd_match['rd'].strip(), sw_id])
                        print('VRF is only available under BGP, Appliance: {0}, vrf name: "{1}", line: "{2}"'.format(
                            do_query('select hostname from appliance where app_id={0};'.format(sw_id), one_row=True),
                            vrf_name, lines[index]
                        ))

                elif re.match(r'\s\S+', lines[index]):
                    # another BGP configuration section other than the last VRF
                    break

        if re.match(r'\S+', lines[index]):
            # outside BGP section
            break

        index = index + 1

    return index


def update_ip_address_int_id(old_int_id, new_int_id):
    qr = do_query(
        'select add_fam_id, address, subnet_id, address_type from pend_ip_address where int_id={0};'.format(old_int_id))
    if len(qr) > 1:
        for ip in qr:
            values = [x for x in ip]
            values.append(new_int_id)
            insert_to_db(
                'ip_address',
                'add_fam_id, address, subnet_id, address_type, int_id',
                values)

        # delete all record for the old_int_id
        do_query('delete from pend_ip_address where int_id={0}'.format(old_int_id))
        return True

    else:
        return False


def get_ip_address_id(ip_add, pending=False):
    if not pending:
        qr = do_query('select address, ip_id from ip_address;')
        for ip in qr:
            if str(ipaddress.ip_interface(ip[0]).ip) == ip_add:
                return ip[1]
    else:
        qr = do_query('select address, ip_id from pend_ip_address;')
        for ip in qr:
            if str(ipaddress.ip_interface(ip[0]).ip) == ip_add:
                return ip[1]

    return ()


def get_int_id(sw_id, int_type, int_number):
    if sw_id and int_type and int_number:
        qr = do_query(
            'select int_id from interface '
            'where app_id={0} and type="{1}" and number="{2}";'.format(sw_id, int_type, int_number), True)
        if qr:
            return qr[0]
        else:
            return ()
    else:
        return ()


def get_vrf_id(vrf_name, sw_id):
    qr = do_query('select vrf_id from vrf where name="{0}" and app_id={1};'.format(vrf_name, sw_id), True)
    if qr:
        return qr[0]
    else:
        return ()


def process_pending_tables():
    """
    Pending processing
    for each pending interface,
        check pending reasons
            vrf name and get vrf_id
        check type, if tunnel
            get source and destination ids
        move int/tun record to normal int table
            update pending_ip_address with the new int_id (if any)
                move the recored to normal ip_address table
            then delete the record (int_record)

        move pending_ips/subnet to normal tables
    :return: None
    """
    repeat = 3
    while repeat != 0:
        qr = do_query(
            'select int_id, type, number, description, mode, member_of, app_id, vrf_name, '
            'tunnel_id, pend_reason, status '
            'from pending_interface')
        if qr:
            for row in qr:
                if row[1] == 'tunnel':
                    vrf_id = do_query(
                        'select * from vrf where name="{0}" and app_id={1}'.format(row[7], row[6]), True)[0]

                    source_int = ()
                    dest_ip_id = None

                    tun = do_query(
                        'select source_int, source_ip, dest_ip '
                        'from pending_tunnel_int '
                        'where tunnel_id={0}'.format(row[8]), True)

                    if tun[0]:
                        # find tunnel source interface id from normal interface table
                        match = re.match(r'(?P<tun_src>(?P<type>[^0-9]+)(?P<number>.+))', tun[0])
                        int_type = match['type']
                        number = match['number']
                        int_id = do_query(
                            'select int_id from interface '
                            'where type="{0}" and number="{1}" and app_id={2}'.format(int_type, number, row[6]),
                            True)

                        if not int_id:
                            continue
                        source_int = (int_id[0], 'source_int')
                    elif tun[1]:
                        # find tunnel source ip id from normal ip_address table
                        tun_src = re.match(r'(?P<tun_src>\d+\.\d+\.\d+\.\d+)', tun[1])['tun_src']
                        ip_id = get_ip_address_id(tun_src)
                        if not ip_id:
                            continue
                        source_int = (ip_id, 'source_ip')
                    if tun[2]:
                        tun_dest = re.match(r'(?P<tun_dest>\d+\.\d+\.\d+\.\d+)', tun[2])['tun_dest']
                        ip_id = get_ip_address_id(tun_dest)
                        if not ip_id:
                            continue
                        dest_ip_id = ip_id

                    # insert a new tunnel record into normal tunnel_int table
                    tun_id = insert_to_db(
                        'tunnel_int', '{0}, dest_ip'.format(source_int[1]), [source_int[0], dest_ip_id])

                    # insert a new interface record into normal interface table
                    int_id = insert_to_db(
                                'interface',
                                'type, number, description, mode, member_of, app_id, vrf_id, tunnel_id, status',
                                [row[1], row[2], row[3], row[4], row[5], row[6], vrf_id, tun_id, row[10]])

                    do_query(
                        'insert into ip_address '
                        '(add_fam_id, address, subnet_id, address_type, int_id) '
                        'select '
                        'add_fam_id, address, subnet_id, address_type, {0} as int_id '
                        'from pend_ip_address where int_id={1};'.format(int_id, row[0]), insert=True)

                    # delete all pending records
                    do_query('delete from pending_tunnel_int where tunnel_id={0}'.format(row[8]))
                    do_query('delete from pending_interface where int_id={0}'.format(row[0]))
                    do_query('delete from pend_ip_address where int_id={0}'.format(row[0]))

        repeat = repeat - 1

    # process pending static routes
    qr = do_query(
        'select next_hop_ip, next_hop_int_type, next_hop_int_number, vrf_name, '
        'subnet_id, ad_distance, name, add_fam_id, app_id, route_id, to_vrf_name '
        'from pend_static_route;')
    if qr:
        for row in qr:
            old_route_id = row[9]
            add_fam_id = row[7]
            sw_id = row[8]
            next_hop_ip_id = None
            if row[0]:
                next_hop_ip_id = insert_ip_subnet(None, add_fam_id, row[0] + '/32', 'next-hop')

            next_hop_int_id = get_int_id(sw_id, row[1], row[2])
            vrf_id = get_vrf_id(row[3], sw_id)
            to_vrf_id = get_vrf_id(row[10], sw_id)
            sub_id = row[4]
            ad = row[5]
            name = row[6]

            insert_to_db('static_route',
                         'app_id, next_hop_ip, next_hop_int, vrf_id, subnet_id, ad_distance, name, '
                         'add_fam_id, to_vrf_id',
                         [sw_id, next_hop_ip_id, next_hop_int_id, vrf_id, sub_id, ad, name, add_fam_id, to_vrf_id])

            do_query('delete from pend_static_route where route_id={0};'.format(old_route_id))


def parse_show_vlan(lines, index, sw_id):
    index = index + 1
    breaks = [r'\s*VLAN\s+Type\s+SAID\s+MTU\s+Parent\s+RingNo\s+BridgeNo\s+Stp\s+BrdgMode\s+Trans1\s+Trans2',
              r'\s*VLAN\s+AREHops\s+STEHops\s+Backup\s+CRF',
              r'\s*Remote\s+SPAN\s+VLANs',
              r'\s*Primary\s+Secondary\s+Type\s+Ports']

    rgx_vlan_id_line = r'\s*(?P<vlan_id>\d+)\s+(?P<name>.+?)\s+' \
                       r'(?P<status>[a-zA-Z]+)\s+(?P<ports>((\w+[0-9\\/]*)(,\s*)?)*)'
    rgx_vlan_id_rem = r'\s+(?P<rem_ports>((\w+[0-9\\/]*)(,\s*)?)*)'

    while index < len(lines):
        match = re.match(rgx_vlan_id_line, lines[index])
        if match:
            vlan_id = int(match['vlan_id'])
            vlan_name = match['name']
            ports = match['ports']
            while True:
                if re.match(rgx_vlan_id_rem, lines[index + 1]):
                    index = index + 1
                    ports = ports + ',' + re.match(rgx_vlan_id_rem, lines[index])['rem_ports']
                else:
                    break
            ports = [x.strip() for x in ports.split(',')]

            insert_to_db('vlan', 'vlan_no, name, app_id, exist', [vlan_id, vlan_name, sw_id, 1])

        index = index + 1
        if re.match(rgx_host, lines[index]) \
                or re.match(breaks[0], lines[index]) \
                or re.match(breaks[1], lines[index]) \
                or re.match(breaks[2], lines[index])\
                or re.match(breaks[3], lines[index]):
            break

    return index


def parse_cdp_nei(lines, index, sw_id):
    index = index + 1

    """Capability Codes: R - Router, T - Trans Bridge, B - Source Route Bridge
                  S - Switch, H - Host, I - IGMP, r - Repeater, P - Phone
    """
    rgx_neig = r'(?P<neighbor>.+?)\s+(?P<local_int>(?P<type1>[^0-9]+?)\s*(?P<number1>[0-9\\/]+))\s+' \
               r'(?P<holdtime>\d+)\s+(?P<capab>([RTBSHIrP]\s)+)\s+' \
               r'(?P<platform>(WS-C.+?)?(CISCO.+?)?(AIR-C.+?)?(IP\s+Phone)?(N\dK-C.+?)?(\d+)?)\s+' \
               r'(?P<remote_int>(?P<type2>[^0-9]+?)\s*(?P<number2>[0-9\\/]+))'
    while True:
        if re.match(rgx_neig, lines[index]):
            match = re.match(rgx_neig, lines[index])
            neigh = match['neighbor']
            local_int = match['local_int']
            remote_int = match['remote_int']
            capabilities = match['capab']
            platform = match['platform']

        elif re.match(rgx_neig, lines[index] + lines[index + 1]):
            match = re.match(rgx_neig, lines[index] + lines[index + 1])
            neigh = match['neighbor']
            local_int = match['local_int']
            remote_int = match['remote_int']
            capabilities = match['capab']
            platform = match['platform']

            index = index + 2
            continue

        elif re.match(rgx_host, lines[index]):
            break

        index = index + 1

    return index


def parse_show_inventory(lines, index, sw_id):
    """
NAME: "Chassis 1 WS-C6509-E", DESCR: "Chassis 1 Cisco Systems, Inc. Catalyst 6500 9-slot Chassis System"
PID: WS-C6509-E        ,                     VID: V07, SN: SMC2122001E

NAME: "0/0/*", DESCR: "Cisco CRS-1 Series Forwarding Processor 40G"
PID: CRS-FP40, VID: V03, SN: SAD1351020H
    [0-9a-zA-Z\\/*=+ ,.-]

    :param lines:
    :param index:
    :param sw_id:
    :return:
    """
    # str1 = 'NAME: "Chassis 1 WS-C6509-E", DESCR: "Chassis 1 Cisco Systems, Inc. Catalyst 6500 9-slot Chassis System"'
    # str2 = 'PID: WS-C6509-E        ,                     VID: V07, SN: SMC2122001E'

    module = {}
    number_of_modules = 0
    line_no = 1

    rgx_att_val_l1 = r'(?P<attribute>\w{3,7}):\s*"?(?P<value>[0-9a-zA-Z\\/*=+ ,.-]+)"?,?'
    rgx_att_val_l2 = r'(?P<attribute>\w{2,7}):\s*(?P<value>[0-9a-zA-Z\\/*=+.-]+),?'

    index = index + 1
    while index < len(lines):
        if re.match(rgx_show_inventory, lines[index]):
            index = index + 1
            continue
        elif re.match(rgx_host, lines[index]):
            break

        line = lines[index]
        match_l1 = re.findall(rgx_att_val_l1, lines[index])
        match_l2 = re.findall(rgx_att_val_l2, lines[index])

        if line_no == 1 and match_l1:
            for att in match_l1:
                if att[0].strip() and att[1].strip():
                    module[att[0].strip()] = att[1].strip()
            line_no = 2

        elif line_no == 2 and match_l2:
            for att in match_l2:
                if att[0].strip() and att[1].strip():
                    module[att[0].strip()] = att[1].strip()
            line_no = 1

        elif not lines[index].strip() and module:
            # print(module)
            insert_to_db('appliance_module', 'name, description, pid, vid, serial, app_id',
                         [
                             module['NAME'] if ('NAME' in module and module['NAME']) else "",
                             module['DESCR'] if ('DESCR' in module and module['DESCR']) else "",
                             module['PID'] if ('PID' in module and module['PID']) else "",
                             module['VID'] if ('VID' in module and module['VID']) else "",
                             module['SN'] if ('SN' in module and module['SN']) else "",
                             sw_id
                         ])
            module = {}
            number_of_modules = number_of_modules + 1
        else:
            if print_ignore_break:
                print('Invalid line ({0}) in show inventory: "{1}"'.format(index, lines[index]))

        index = index + 1
    print('Number of number_of_modules found: {0}'.format(number_of_modules))
    return index


def save_inventory(out_file):
    """
    name, description, pid, vid, serial
    :param out_file:
    :return:
    """
    qr = do_query('select appliance.hostname, appliance_module.name, appliance_module.description, '
                  'appliance_module.pid, appliance_module.vid, appliance_module.serial '
                  'from appliance_module '
                  'join appliance on appliance.app_id = appliance_module.app_id')
    if qr:
        with open(out_file, 'w+') as f:
            f.write('Hostname,Name,Description,PID,VID,Serial\n')
            for row in qr:
                f.write('{0},{1},{2},{3},{4},{5}\n'.format(row[0], row[1], row[2], row[3], row[4], row[5]))


def parse_log_file(log_file, site_id):
    ios_ver = 'ios'

    hostname = ''
    sw_id = None
    vrfs = []
    vrf_rt_to_name = {}
    with open(log_file, "r+b") as f:
        data = f.read()
        # For converting from UTF-8 BOM to UTF-8
        data = data.decode("utf-8-sig")
        lines = data.splitlines()

        i = 0
        while i < len(lines):

            if not hostname and re.match(rgx_host, lines[i]):
                # get hostname from the CLI prompt

                hostname = re.match(rgx_host, lines[i])['hostname']

                sw_id = insert_to_db('appliance', 'hostname, site_id', [hostname, site_id])
                match = re.match(r'\s*((\w:)?(.+?[\\/]{1,2})*)(?P<name>(?P<num>\d*)\s*(.+?)(\.\w+))', log_file)

                insert_to_db('log_file',
                             'filename, file_text, app_id, importance',
                             [match['name'], str(data), sw_id, int(match['num']) if match['num'] else 0])
                if data:
                    del data

            if re.match(rgx_show_run_sec, lines[i]):
                i = i + 1

                while i < len(lines):
                    if not hostname and re.match(r'\s*hostname\s*(?P<hostname>.+)', lines[i]):
                        # match hostname from hostname in the running-config

                        hostname = re.match(r'\s*hostname\s*(?P<hostname>.+)', lines[i])['hostname']

                        sw_id = insert_to_db('appliance', 'hostname, site_id', [hostname, site_id])
                        match = re.match(r'\s*((\w:)?(.+?[\\/]{1,2})*)(?P<name>(?P<num>\d*)\s*(.+?)(\.\w+))', log_file)

                        insert_to_db('log_file',
                                     'filename, file_text, app_id, importance',
                                     [match['name'], str(data), sw_id, int(match['num']) if match['num'] else 0])
                        if data:
                            del data

                        i = i + 1

                    if re.match(rgx_ios_ip_vrf, lines[i]) and not re.match(rgx_vrf_fwd, lines[i]):
                        # add_family 'IPv4' id = 1
                        # imp_exp = [exports, imports]

                        vrf_name = re.match(rgx_ios_ip_vrf, lines[i])['vrf_name']
                        rd, imp_exp, description, i = parse_vrf(lines, i, vrf_name, vrf_rt_to_name)
                        if sw_id:
                            insert_vrf_to_db(sw_id, vrf_name, rd, imp_exp, description)
                        else:
                            vrfs.append([sw_id, vrf_name, rd, imp_exp, description])

                    elif re.match(rgx_ios_vrf_def, lines[i]):
                        # imp_exp = [exports, imports, exportsv6, importsv6]
                        vrf_name = re.match(rgx_ios_vrf_def, lines[i])['vrf_name']
                        rd, imp_exp, description, i = parse_vrf(lines, i, vrf_name, vrf_rt_to_name, vrf_def=True)
                        if sw_id:
                            insert_vrf_to_db(sw_id, vrf_name, rd, imp_exp, description)
                        else:
                            vrfs.append([sw_id, vrf_name, rd, imp_exp, description])

                    elif re.match(rgx_xr_vrf, lines[i]):
                        # IOS XR VRF configuration doesn't include RD
                        # RD configured under BGP VRF section "PE_to_CE"
                        ios_ver = 'xr'
                        vrf_name = re.match(rgx_xr_vrf, lines[i])['vrf_name']
                        imp_exp, description, i = parse_vrf(lines, i, vrf_name, vrf_rt_to_name, crs_asr=True)

                        # rd will be replace with ""
                        # after parsing BGP section, rd should be updated

                        if sw_id:
                            insert_vrf_to_db(sw_id, vrf_name, "", imp_exp, description)
                        else:
                            vrfs.append([sw_id, vrf_name, "", imp_exp, description])

                    elif re.match(r'\s*vlan\s+(?P<id>\d+(-\d+)?(,\d+(-\d+)?)*)', lines[i]):
                        i = parse_vlan(lines, i, sw_id)

                    elif re.match(rgx_ios_static_route, lines[i]):
                        i = parse_static_route(lines, i, sw_id)

                    elif re.match(rgx_xr_router_static, lines[i]):
                        i = parse_xr_static_route(lines, i, sw_id)

                    elif re.match(rgx_interface, lines[i]):
                        i = parse_all_interface_type(sw_id, ios_ver, lines, i)

                    elif re.match(r'router\s+bgp\s+(?P<as>\d+([.]\d+)?)', lines[i]):
                        if ios_ver == 'xr':
                            i = parse_xr_router_bgp(lines, i, sw_id)

                    elif re.match(rgx_host, lines[i]) or re.match(r'end\s*$', lines[i]):
                        break

                    else:
                        i = i + 1

                if len(vrfs) > 0:
                    for vrf in vrfs:
                        insert_vrf_to_db(vrf[0], vrf[1], vrf[2], vrf[3], vrf[4])
                    vrfs = []

            elif re.match(rgx_show_vlan_sec, lines[i]):
                i = parse_show_vlan(lines, i, sw_id)

            elif re.match(rgx_cdp_nei_sec, lines[i]):
                i = parse_cdp_nei(lines, i, sw_id)

            elif re.match(rgx_show_inventory, lines[i]):
                i = parse_show_inventory(lines, i, sw_id)

            else:
                # print('Unknown line: ', lines[i])
                i = i + 1
        if not hostname:
            print("Hostname couldn't be found for app_id: {0}".format(sw_id))


# @profile
def main():
    path = 'configs\\RC\\'
    file_extension = '.log'

    db_dump_file = 'db_dump.sql'

    global db_con, db_cur
    db_con, db_cur = create_db()

    vrf_wb = Workbook()
    vlan_wb = Workbook()
    prepare_workbooks_styles((vrf_wb, vlan_wb))

    file_list = get_files_from_path(path, file_extension)

    if file_list:
        print('Number of files: {0}'.format(len(file_list)))

        file_list.sort(key=natural_sort)
        for file_name in file_list:
            print(file_name)
            site_id = insert_to_db('site', 'name', re.match(r'\s*\d*\s*(?P<name>.+)', file_name)['name'])
            parse_log_file(path + file_name + file_extension, site_id)

        process_pending_tables()

        out_file_name = 'mpls_view.xlsx'
        save_vrfs_to_excel(vrf_wb)

        headers = ['VRF Name', 'Export RT', 'Interface', 'IP Address (Primary/VIP)', 'Interface Description']
        save_int_to_excel(vlan_wb, headers)
        save_inventory(path + 'Inventory.csv')

        vrf_wb.save(path + out_file_name)
        vlan_wb.save(path + 'VRF_VLANs.xlsx')

        db_close_dump(path + db_dump_file)

    else:
        print("No file found with extension {0} is found!!".format(file_extension))


if __name__ == '__main__':
    start = time.process_time()
    main()
    # parse_show_inventory()
    print('Time elapsed {:3.2f} seconds'.format(time.process_time() - start))
